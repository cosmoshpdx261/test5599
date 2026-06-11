import streamlit as st
import pandas as pd
import numpy as np
import random
import jpholiday
import datetime
import calendar
import os
import json

# 🌟【ここから追加】病院独自の休日をシステム全体に認識させる魔法のコード🌟
original_jpholiday_is_holiday = jpholiday.is_holiday

def custom_is_holiday(date_obj):
    # 1. カレンダー通りの祝日ならTrue
    if original_jpholiday_is_holiday(date_obj):
        return True
    
    # 2. 病院独自の特別休暇（お盆・年末年始）ならTrue
    m = date_obj.month
    d = date_obj.day
    if m == 8 and d in [13, 14, 15]:
        return True
    if m == 12 and d in [30, 31]:
        return True
    if m == 1 and d in [1, 2, 3]:
        return True
        
    return False

# システム標準の祝日判定を、コスモス病院専用の判定にすり替える（上書き）
jpholiday.is_holiday = custom_is_holiday
# 🌟【ここまで追加】🌟


# ページ設定（ワイドモード）
st.set_page_config(
    page_title="コスモス外来・OPEシフト作成システム",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==========================================
# 0. 設定・セッション状態の初期化
# ==========================================

ADMIN_PASSWORD = "okuzaki"
DEADLINE_DAY = 15 # ★希望提出の締め切り日

if 'generated_df' not in st.session_state:
    st.session_state['generated_df'] = None
if 'meta_color_df' not in st.session_state:
    st.session_state['meta_color_df'] = None
if 'e_schedule_df' not in st.session_state:
    st.session_state['e_schedule_df'] = None
if 'meta_e_color_df' not in st.session_state:
    st.session_state['meta_e_color_df'] = None
if 'wishes' not in st.session_state:
    st.session_state['wishes'] = []

if 'schedule_history' not in st.session_state:
    st.session_state['schedule_history'] = []
if 'run_generation' not in st.session_state:
    st.session_state['run_generation'] = None
if 'admin_logged_in' not in st.session_state:
    st.session_state['admin_logged_in'] = False
if 'staff_logged_in_id' not in st.session_state:
    st.session_state['staff_logged_in_id'] = None
if 'previous_app_mode' not in st.session_state:
    st.session_state['previous_app_mode'] = "👤 スタッフ入力画面"
if 'editor_reset_key' not in st.session_state:
    st.session_state['editor_reset_key'] = 0
if 'current_view_month' not in st.session_state:
    st.session_state['current_view_month'] = None
if 'schedule_locked' not in st.session_state:
    st.session_state['schedule_locked'] = False

# 定数定義
SHIFT_TYPES = ['日', 'OP日', '内日', 'A', 'B', '当直', 'OFF', '遅', '半/遅', '年/遅', '年休', '特休', '夏休', '/年', '年/', '/半', '半/', '年/半', '/特', '特/', '/夏', '夏/', '代休', '/代休', '研修', '/研', '研/', '健/年', '健/半', '健/', '---', '介護休暇', '病欠', '産休', '育休', '学', 'P4', 'P5', 'P8', 'P11', 'P13', 'P17', 'P70', '内P7', '内P9', '内P16']
ROLES = ['師長', '副師長', '看護師', 'CW', 'CE', '事務'] 
GROUPS = ['外来', 'オペ室', '内視鏡'] 

STAFF_FILE = 'staff_data.csv'
WISHES_FILE = 'wishes_data.csv'
REMARKS_FILE = 'remarks_data.csv' 
EVENTS_FILE = 'events_data.csv' 
COMMITTEES_FILE = 'committees_data.csv' 
SCHEDULE_DATA_FILE = 'schedule_data.json'
CLOSED_DAYS_FILE = 'closed_days_data.csv' 

# ★Aは休日カウント対象から除外
HOLIDAY_WORK_SHIFTS = ['日', '遅', '半/遅', '年/遅'] 
WORK_SHIFTS = ['日', 'OP日', '内日', 'A', 'B', '当直', '遅', '半/遅', '年/遅', '/半', '半/', '/特', '特/', '/夏', '夏/', '研修', '/研', '研/', '学', 'P4', 'P5', 'P8', 'P11', 'P13', 'P17', 'P70', '内P7', '内P9', '内P16']
E_CAPABLE_SHIFTS = ['日', 'OP日', '内日', 'B', '半/遅', '年/遅']
REST_AND_A_SHIFTS = ['OFF', '年休', '特休', '夏休', '/年', '年/', '/半', '半/', '年/半', '/特', '特/', '/夏', '夏/', '健/年', '健/半', '健/', 'A', '介護休暇', '病欠', '産休', '育休', '代休', '/代休']
# ==========================================
# 1. データ生成・読み込み・保存関数
# ==========================================

def load_all_schedules():
    if os.path.exists(SCHEDULE_DATA_FILE):
        try:
            with open(SCHEDULE_DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return {}

def save_all_schedules(data):
    with open(SCHEDULE_DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)

def get_monthly_schedule(year, month):
    data = load_all_schedules()
    key = f"{year}-{month:02d}"
    if key in data:
        try:
            gen_df = pd.DataFrame(data[key]['gen']).replace({None: np.nan, '半遅': '半/遅'})
            gen_df.index = [int(x) if str(x).isdigit() else x for x in gen_df.index]
            
            col_df = pd.DataFrame(data[key]['col'])
            col_df.index = [int(x) if str(x).isdigit() else x for x in col_df.index]
            
            e_df = pd.DataFrame(data[key]['e_gen']).replace({None: np.nan, '半遅': '半/遅'})
            e_col_df = pd.DataFrame(data[key]['e_col'])
            
            return gen_df, col_df, e_df, e_col_df
        except:
            pass
    return None, None, None, None

def save_monthly_schedule(year, month, gen_df, col_df, e_df, e_col_df):
    data = load_all_schedules()
    key = f"{year}-{month:02d}"
    
    data[key] = {
        'gen': gen_df.where(pd.notnull(gen_df), None).to_dict(),
        'col': col_df.to_dict(),
        'e_gen': e_df.where(pd.notnull(e_df), None).to_dict(),
        'e_col': e_col_df.to_dict()
    }
    save_all_schedules(data)

def persist_current_schedule(year, month):
    if st.session_state['generated_df'] is not None:
        save_monthly_schedule(
            year, 
            month, 
            st.session_state['generated_df'], 
            st.session_state['meta_color_df'], 
            st.session_state['e_schedule_df'], 
            st.session_state['meta_e_color_df']
        )

def apply_rollover(gen_df, col_df, target_year, target_month, sids, date_cols):
    if target_month == 1:
        p_year = target_year - 1
        p_month = 12
    else:
        p_year = target_year
        p_month = target_month - 1
        
    p_gen, p_col, _, _ = get_monthly_schedule(p_year, p_month)
    
    if p_gen is not None:
        num_days_p = calendar.monthrange(p_year, p_month)[1]
        last_day = f"{p_year}-{p_month:02d}-{num_days_p:02d}"
        
        for s in sids:
            if s in p_gen.index and last_day in p_gen.columns:
                val_last = p_gen.at[s, last_day]
                
                if val_last == 'A':
                    day1 = f"{target_year}-{target_month:02d}-01"
                    day2 = f"{target_year}-{target_month:02d}-02"
                    if day1 in date_cols and col_df.at[s, day1] not in ['green', 'red']:
                        gen_df.at[s, day1] = 'B'
                        col_df.at[s, day1] = 'green'
                    if day2 in date_cols and col_df.at[s, day2] not in ['green', 'red']:
                        gen_df.at[s, day2] = 'OFF'
                        col_df.at[s, day2] = 'green'
                        
                elif val_last == 'B':
                    day1 = f"{target_year}-{target_month:02d}-01"
                    if day1 in date_cols and col_df.at[s, day1] not in ['green', 'red']:
                        gen_df.at[s, day1] = 'OFF'
                        col_df.at[s, day1] = 'green'
                        
                elif val_last in ['遅', '半/遅', '年/遅']:
                    last_date_obj = datetime.date(p_year, p_month, num_days_p)
                    if (val_last == '遅' and last_date_obj.weekday() == 5 and not jpholiday.is_holiday(last_date_obj)) or \
                       (val_last in ['半/遅', '年/遅'] and (last_date_obj.weekday() == 6 or jpholiday.is_holiday(last_date_obj))):
                        day1 = f"{target_year}-{target_month:02d}-01"
                        if day1 in date_cols and col_df.at[s, day1] not in ['green', 'red']:
                            gen_df.at[s, day1] = 'OFF'
                            col_df.at[s, day1] = 'green'

def build_draft_schedule(year, month, staff_df_local, wishes_local, committees_local):
    dates_local = get_dates_in_month(year, month)
    cols_local = [d.strftime('%Y-%m-%d') for d in dates_local]
    sids_local = staff_df_local["職員番号"].values
    
    gen_df = pd.DataFrame(np.nan, index=sids_local, columns=cols_local)
    col_df = pd.DataFrame("black", index=sids_local, columns=cols_local)
    e_df = pd.DataFrame(np.nan, index=['E'], columns=cols_local)
    e_col_df = pd.DataFrame("black", index=['E'], columns=cols_local)
    
    staff_info = staff_df_local.set_index("職員番号").to_dict(orient="index")
    
    apply_rollover(gen_df, col_df, year, month, sids_local, cols_local)
    
    for w in wishes_local:
        sid = w['staff_id']
        if sid in gen_df.index:
            d_s = w['date']
            if d_s in cols_local:
                if col_df.at[sid, d_s] != 'green': 
                    gen_df.at[sid, d_s] = w['shift']
                    col_df.at[sid, d_s] = w['type'] 
                    
    for c in committees_local:
        if c['year'] == year and c['month'] == month:
            d_s = c['date']
            if d_s in cols_local:
                for sid in c['members']:
                    if sid in gen_df.index:
                        if col_df.at[sid, d_s] != 'green':
                            wished_off = any(w['staff_id'] == sid and w['date'] == d_s and w['shift'] in ['OFF', '年休', '特休', '夏休', '/年', '年/', '/半', '半/', '年/半', '/特', '特/', '/夏', '夏/', '介護休暇', '病欠', '産休', '育休', '代休', '/代休'] for w in wishes_local)
                            if not wished_off and pd.isna(gen_df.at[sid, d_s]):
                                # 🌟【修正】パートさんの基本勤務を優先する
                                b_s = staff_info[sid].get('基本勤務', '日')
                                if pd.isna(b_s) or b_s == '': b_s = '日'
                                
                                if b_s != '日':
                                    def_c_shift = b_s
                                else:
                                    def_c_shift = '日'
                                    if staff_info[sid]['所属'] == 'オペ室': def_c_shift = 'OP日'
                                    elif staff_info[sid]['所属'] == '内視鏡': def_c_shift = '内日'
                                    
                                gen_df.at[sid, d_s] = def_c_shift
                                
    return gen_df, col_df, e_df, e_col_df

def create_default_staff_data():
    return pd.DataFrame(columns=["職員番号", "名前", "所属", "役職", "パスワード", "表示順", "夜勤上限", "休日日勤上限", "遅出半遅上限", "待機上限", "DM", "リブレ", "学生", "基本勤務"])

def load_or_create_staff_data():
    if os.path.exists(STAFF_FILE):
        try:
            df = pd.read_csv(STAFF_FILE)
            if "ID" in df.columns and "職員番号" not in df.columns:
                df.rename(columns={"ID": "職員番号"}, inplace=True)
            if "パスワード" not in df.columns: df["パスワード"] = "1234"
            if "表示順" not in df.columns: df["表示順"] = 999
            if "夜勤上限" not in df.columns: df["夜勤上限"] = 4
            
            # ★ 古い「休日上限」を「休日日勤上限」に引き継ぎ、新しく「遅出半遅上限」を作る
            if "休日上限" in df.columns and "休日日勤上限" not in df.columns:
                df.rename(columns={"休日上限": "休日日勤上限"}, inplace=True)
            if "休日日勤上限" not in df.columns: df["休日日勤上限"] = "無し"
            if "遅出半遅上限" not in df.columns: df["遅出半遅上限"] = "無し"
            
            if "所属" not in df.columns: df["所属"] = "外来" 
            if "待機上限" not in df.columns: df["待機上限"] = 0
            if "役職" in df.columns: df.loc[df["役職"] == "ケアワーカー", "役職"] = "CW"
            if "DM" not in df.columns: df["DM"] = False
            if "リブレ" not in df.columns: df["リブレ"] = False
            if "リエゾン" in df.columns: df.drop(columns=["リエゾン"], inplace=True)
            if "学生" not in df.columns: df["学生"] = False
            # 🌟【追加】今いるスタッフ全員の初期値を「日」にする
            if "基本勤務" not in df.columns: df["基本勤務"] = "日"
                
            df.to_csv(STAFF_FILE, index=False)
            
            if not df.empty:
                df["パスワード"] = df["パスワード"].astype(str)
                df["職員番号"] = df["職員番号"].astype(int)
                df["表示順"] = df["表示順"].astype(int)
                df["夜勤上限"] = df["夜勤上限"].astype(int)
                df["休日日勤上限"] = df["休日日勤上限"].astype(str)
                df["遅出半遅上限"] = df["遅出半遅上限"].astype(str)
                df["所属"] = df["所属"].astype(str)
                df["待機上限"] = df["待機上限"].fillna(0).astype(int)
                df["DM"] = df["DM"].astype(bool)
                df["リブレ"] = df["リブレ"].astype(bool)
                df["学生"] = df["学生"].astype(bool)
                # 🌟【追加】文字データとして読み込む
                df["基本勤務"] = df["基本勤務"].astype(str)
            return df
        except pd.errors.EmptyDataError:
            pass
    df = create_default_staff_data()
    df.to_csv(STAFF_FILE, index=False)
    return df
              
def save_staff_data(df):
    df.to_csv(STAFF_FILE, index=False)

def load_wishes_data():
    if os.path.exists(WISHES_FILE):
        try: 
            df = pd.read_csv(WISHES_FILE)
            if 'shift' in df.columns:
                df['shift'] = df['shift'].replace({'半遅': '半/遅'})
            return df.to_dict('records')
        except pd.errors.EmptyDataError: 
            return []
    return []

def save_wish_data(wish_list):
    pd.DataFrame(wish_list).to_csv(WISHES_FILE, index=False)

def load_remarks_data():
    if os.path.exists(REMARKS_FILE):
        try: 
            return pd.read_csv(REMARKS_FILE).to_dict('records')
        except pd.errors.EmptyDataError: 
            return []
    return []

def save_remarks_data(remarks_list):
    pd.DataFrame(remarks_list).to_csv(REMARKS_FILE, index=False)

def load_events_data():
    if os.path.exists(EVENTS_FILE):
        try: 
            return pd.read_csv(EVENTS_FILE).to_dict('records')
        except pd.errors.EmptyDataError: 
            return []
    return []

def save_events_data(events_list):
    pd.DataFrame(events_list).to_csv(EVENTS_FILE, index=False)

def load_committees_data():
    if os.path.exists(COMMITTEES_FILE):
        try: 
            df = pd.read_csv(COMMITTEES_FILE)
            df['members'] = df['members'].apply(lambda x: json.loads(x) if isinstance(x, str) else [])
            return df.to_dict('records')
        except pd.errors.EmptyDataError: 
            return []
    return []

def save_committees_data(committees_list):
    df = pd.DataFrame(committees_list)
    if not df.empty and 'members' in df.columns:
        df['members'] = df['members'].apply(json.dumps)
    df.to_csv(COMMITTEES_FILE, index=False)

def load_closed_days_data():
    if os.path.exists(CLOSED_DAYS_FILE):
        try: 
            return pd.read_csv(CLOSED_DAYS_FILE).to_dict('records')
        except pd.errors.EmptyDataError: 
            return []
    return []

def save_closed_days_data(closed_list):
    pd.DataFrame(closed_list).to_csv(CLOSED_DAYS_FILE, index=False)

def get_dates_in_month(year, month):
    num_days = calendar.monthrange(year, month)[1]
    return [datetime.date(year, month, day) for day in range(1, num_days + 1)]

def is_holiday_jp(date_obj):
    return date_obj.weekday() >= 5 or jpholiday.is_holiday(date_obj)

def get_allowed_shifts(role):
    special_leaves = ['介護休暇', '病欠', '産休', '育休']
    base_shifts = [s for s in SHIFT_TYPES if s not in ['---', '健/年', '健/半', '健/', '学'] + special_leaves]
    if role == "事務": return ['OFF', '年休', '特休', '夏休', '/年', '年/', '/半', '半/', '年/半', '/特', '特/', '/夏', '夏/', '代休', '/代休', '研修', '/研', '研/']
    elif role == "CW": return ['日', 'OP日', '内日', '遅', 'OFF', '年休', '特休', '夏休', '/年', '年/', '/半', '半/', '年/半', '/特', '特/', '/夏', '夏/', '代休', '/代休', '研修', '/研', '研/', 'P4', 'P5', 'P8', 'P11', 'P13', 'P17', 'P70', '内P7', '内P9', '内P16']
    elif role == "師長": return ['日', 'OP日', '内日', '当直', 'OFF', '年休', '特休', '夏休', '/年', '年/', '/半', '半/', '年/半', '/特', '特/', '/夏', '夏/', '代休', '/代休', '研修', '/研', '研/']
    elif role == "CE": return ['日', 'OP日', '内日', '遅', '半/遅', '年/遅', 'OFF', '年休', '特休', '夏休', '/年', '年/', '/半', '半/', '年/半', '/特', '特/', '/夏', '夏/', '代休', '/代休', '研修', '/研', '研/']
    elif role == "看護師": return ['日', 'OP日', '内日', 'A', 'B', '遅', '半/遅', '年/遅', 'OFF', '年休', '特休', '夏休', '/年', '年/', '/半', '半/', '年/半', '/特', '特/', '/夏', '夏/', '代休', '/代休', '研修', '/研', '研/', 'P4', 'P5', 'P8', 'P11', 'P13', 'P17', 'P70', '内P7', '内P9', '内P16']
    else: return base_shifts

def check_shift_rule(role, date_obj, shift):
    if shift == "-": return None
    is_hol = is_holiday_jp(date_obj)
    d_str = date_obj.strftime('%d日')
    if role == "事務" and is_hol and shift not in ['OFF', '年休', '特休', '夏休', '/年', '年/', '/半', '半/', '年/半', '/特', '特/', '/夏', '夏/', '代休', '/代休']:
        return f"🚫 {d_str}: 事務職は土日祝に勤務希望を出せません。"
    if is_hol and shift in ['OP日', '内日', '内P7', '内P9', '内P16']:
        return f"🚫 {d_str}: 土日祝日に専門部署の記号（{shift}）は選択できません（「日」などを選択してください）。"
    allowed = get_allowed_shifts(role)
    if shift not in allowed: return f"🚫 {d_str}: {role}はこのシフト（{shift}）を選択できません。"
    return None

def restore_staff_inputs(my_wishes):
    red_list = sorted([w for w in my_wishes if w['type'] == 'red'], key=lambda x: x['date'])
    blue_list = sorted([w for w in my_wishes if w['type'] == 'blue'], key=lambda x: x['date'])
    red_sets = []
    used_indices = set()
    for i in range(len(red_list)):
        if i in used_indices: continue
        start_date_str = red_list[i]['date']
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        current_set = {"start": start_date, "shifts": ["-", "-", "-"]}
        current_set["shifts"][0] = red_list[i]['shift']
        used_indices.add(i)
        day2_date = start_date + datetime.timedelta(days=1)
        day3_date = start_date + datetime.timedelta(days=2)
        for j in range(len(red_list)):
            if j in used_indices: continue
            d_str = red_list[j]['date']
            d_obj = datetime.datetime.strptime(d_str, '%Y-%m-%d').date()
            if d_obj == day2_date:
                current_set["shifts"][1] = red_list[j]['shift']
                used_indices.add(j)
            elif d_obj == day3_date:
                current_set["shifts"][2] = red_list[j]['shift']
                used_indices.add(j)
        red_sets.append(current_set)
        if len(red_sets) >= 2: break
    blue_sets = []
    for item in blue_list[:2]:
        d_obj = datetime.datetime.strptime(item['date'], '%Y-%m-%d').date()
        blue_sets.append({"date": d_obj, "shift": item['shift']})
    return red_sets, blue_sets

def can_work_holiday_day(staff_id, current_count, staff_info_dict):
    limit_str = str(staff_info_dict[staff_id].get('休日日勤上限', '無し'))
    if limit_str == '無し': return True
    return current_count < int(limit_str)

def can_work_late_shift(staff_id, current_count, staff_info_dict):
    limit_str = str(staff_info_dict[staff_id].get('遅出半遅上限', '無し'))
    if limit_str == '無し': return True
    return current_count < int(limit_str)

def is_consecutive_holiday_work(schedule, sid, date_idx, dates):
    prev_work = False
    if date_idx > 0:
        prev_d = dates[date_idx - 1]
        if is_holiday_jp(prev_d):
            prev_s = prev_d.strftime('%Y-%m-%d')
            if schedule.at[sid, prev_s] in HOLIDAY_WORK_SHIFTS:
                prev_work = True
    next_work = False
    if date_idx + 1 < len(dates):
        next_d = dates[date_idx + 1]
        if is_holiday_jp(next_d):
            next_s = next_d.strftime('%Y-%m-%d')
            if schedule.at[sid, next_s] in HOLIDAY_WORK_SHIFTS:
                next_work = True
    return prev_work or next_work

# ==========================================
# 2. 確認ダイアログ関数
# ==========================================
@st.dialog("登録確認")
def confirm_registration(new_id, new_name, new_group, new_role, new_order, new_night_limit, new_hol_day_limit, new_late_limit, new_e_limit, new_student, current_df):
    st.write("以下の内容で新規登録しますか？")
    st.info(f"職員番号: {new_id}\n氏名: {new_name}\n所属: {new_group}\n役職: {new_role}\n表示順: {new_order}\n夜勤上限: {new_night_limit}\n休日日勤上限: {new_hol_day_limit}\n遅出半遅上限: {new_late_limit}\n待機(E)上限: {new_e_limit}\n学生: {new_student}")
    if st.button("登録実行", type="primary"):
        new_record = pd.DataFrame([{
            "職員番号": int(new_id), "名前": new_name, "所属": new_group, "役職": new_role, 
            "パスワード": "1234", "表示順": int(new_order), 
            "夜勤上限": int(new_night_limit), "休日日勤上限": str(new_hol_day_limit), "遅出半遅上限": str(new_late_limit), "待機上限": int(new_e_limit),
            "DM": False, "リブレ": False, "学生": new_student
        }])
        updated_df = pd.concat([current_df, new_record], ignore_index=True)
        save_staff_data(updated_df)
        st.success(f"{new_name} さんを登録しました")
        st.rerun()

@st.dialog("修正確認")
def confirm_update(target_id, new_name, new_group, new_role, new_order, new_night_limit, new_hol_day_limit, new_late_limit, current_df):
    st.write("以下の内容で情報を更新しますか？")
    st.info(f"職員番号: {target_id}\n氏名: {new_name}\n所属: {new_group}\n役職: {new_role}\n表示順: {new_order}\n夜勤上限: {new_night_limit}\n休日日勤上限: {new_hol_day_limit}\n遅出半遅上限: {new_late_limit}")
    if st.button("修正を実行", type="primary"):
        current_df.loc[current_df["職員番号"] == target_id, ["名前", "所属", "役職", "表示順", "夜勤上限", "休日日勤上限", "遅出半遅上限"]] = [new_name, new_group, new_role, int(new_order), int(new_night_limit), str(new_hol_day_limit), str(new_late_limit)]
        save_staff_data(current_df)
        st.success("情報を更新しました")
        st.rerun()

@st.dialog("削除確認")
def confirm_delete(target_id, target_name, current_df):
    st.write("本当に以下の職員を削除しますか？")
    st.error(f"職員番号: {target_id}\n\n氏名: {target_name}")
    if st.button("削除実行", type="primary"):
        save_staff_data(current_df[current_df["職員番号"] != target_id])
        st.success("削除しました")
        st.rerun()

@st.dialog("登録完了")
def show_registration_success(valid_records, remark_text=""):
    sorted_records = sorted(valid_records, key=lambda x: x['date'])
    msg_lines = []
    for r in sorted_records:
        d_obj = datetime.datetime.strptime(r['date'], '%Y-%m-%d')
        date_str = f"{d_obj.month}月{d_obj.day}日"
        msg_lines.append(f"{date_str}　{r['shift']}")
    if msg_lines:
        st.markdown("\n\n".join(msg_lines))
        st.write("")
        st.write("で登録されました")
    else: 
        st.write("希望シフトは登録されていません（全てクリアされました）。")
    if remark_text:
        st.write("---")
        st.write("**【備考】** も登録されました：")
        st.write(remark_text)
    if st.button("OK", type="primary"): 
        st.rerun()

@st.dialog("⚠️ 変更の確認")
def confirm_sensitive_changes(changes, new_gen_df, new_col_df, e_changed, new_e_df, new_e_col, year, month):
    st.write("以下の**希望（赤・青）**が変更されています。")
    st.write("このまま変更を反映しますか？")
    for c in changes:
        type_str = "赤希望" if c['color'] == 'red' else "青希望"
        st.error(f"{c['name']}さんの {c['date_label']} : {type_str} ({c['old']} → {c['new']})")
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("変更する (反映)", type="primary"):
            hist = st.session_state.get('schedule_history', [])
            hist.append({
                'gen': st.session_state['generated_df'].copy(deep=True),
                'col': st.session_state['meta_color_df'].copy(deep=True),
                'e_gen': st.session_state['e_schedule_df'].copy(deep=True),
                'e_col': st.session_state['meta_e_color_df'].copy(deep=True)
            })
            if len(hist) > 4: hist.pop(0)
            st.session_state['schedule_history'] = hist
            
            st.session_state['generated_df'] = new_gen_df
            st.session_state['meta_color_df'] = new_col_df
            if e_changed:
                st.session_state['e_schedule_df'] = new_e_df
                st.session_state['meta_e_color_df'] = new_e_col
                        
            persist_current_schedule(year, month)
            st.success("✅ シフトを確定保存しました！")
            st.rerun()
            
    with col2:
        if st.button("キャンセル (元に戻す)"):
            st.session_state['editor_reset_key'] += 1
            st.rerun()

@st.dialog("自動生成の確認")
def confirm_generation_dialog(mode):
    if mode == "keep":
        st.write("手動で確定したシフト（🟢）と下書き（🔴🔵🟡）を保持し、残りの空き枠を自動的に埋めます。")
        st.write("よろしいですか？")
    else:
        st.error("⚠️ 手動で確定したシフト（🟢）をすべてリセットし、初期の下書き状態に戻します。")
        st.write("よろしいですか？")
        
    col1, col2 = st.columns(2)
    with col1:
        if st.button("実行する", type="primary"):
            st.session_state['run_generation'] = mode
            st.rerun()
    with col2:
        if st.button("キャンセル"): 
            st.rerun()
            
@st.dialog("🎯 ピンポイント穴埋めアシスト")
def pinpoint_fill_dialog(date_strs_local, col_map_local, staff_df_local, wishes_local, year, month):
    st.write("指定した日の不足枠に、現在OFFの人を1名自動で補充します。（他の日は動きません）")
    hole_date = st.selectbox("📅 補充する日付", options=date_strs_local, format_func=lambda x: col_map_local[x])
    hole_shift = st.selectbox("📝 補充するシフト", options=["日", "遅", "半/遅", "年/遅", "A", "B", "OP日", "内日"])

    if st.button("🚀 補充を実行する", type="primary"):
        current_sch = st.session_state['generated_df']
        staff_info = staff_df_local.set_index("職員番号").to_dict(orient="index")
        hole_d_obj = datetime.datetime.strptime(hole_date, '%Y-%m-%d').date()
        is_hole_hol = is_holiday_jp(hole_d_obj)

        holiday_day_counts = {s: 0 for s in staff_df_local["職員番号"].values}
        late_counts = {s: 0 for s in staff_df_local["職員番号"].values}
        half_late_counts = {s: 0 for s in staff_df_local["職員番号"].values}
        night_counts = {s: 0 for s in staff_df_local["職員番号"].values}

        for s in current_sch.index:
            for d_str in date_strs_local:
                val = current_sch.at[s, d_str]
                if pd.notna(val):
                    d_obj = datetime.datetime.strptime(d_str, '%Y-%m-%d').date()
                    if val == '日' and is_holiday_jp(d_obj): holiday_day_counts[s] += 1
                    if val == '遅': late_counts[s] += 1
                    if val in ['半/遅', '年/遅']: half_late_counts[s] += 1
                    if val in ['A', '当直']: night_counts[s] += 1

        cands = []
        for sid in staff_df_local["職員番号"].values:
            if sid not in current_sch.index: continue
            current_val = current_sch.at[sid, hole_date]

            if pd.isna(current_val) or current_val == 'OFF' or current_val == '---':
                wished = [w for w in wishes_local if w['staff_id'] == sid and w['date'] == hole_date]
                if any(w['shift'] in ['OFF', '年休', '特休', '夏休', '産休', '育休', '病欠', '介護休暇', '代休', '/代休'] for w in wished):
                    continue

                role = staff_info[sid]['役職']
                if hole_shift not in get_allowed_shifts(role): continue

                valid = True
                if is_hole_hol and hole_shift == '日':
                    if not can_work_holiday_day(sid, holiday_day_counts[sid], staff_info): valid = False
                if hole_shift in ['遅', '半/遅', '年/遅']:
                    if not can_work_late_shift(sid, late_counts[sid] + half_late_counts[sid], staff_info): valid = False
                if hole_shift in ['A', '当直']:
                    night_limit = staff_info[sid].get('夜勤上限', 4)
                    if night_counts[sid] >= night_limit: valid = False

                if valid:
                    cands.append(sid)

        if cands:
            chosen_sid = random.choice(cands)
            chosen_name = staff_info[chosen_sid]['名前']

            hist = st.session_state.get('schedule_history', [])
            hist.append({
                'gen': current_sch.copy(deep=True),
                'col': st.session_state['meta_color_df'].copy(deep=True),
                'e_gen': st.session_state['e_schedule_df'].copy(deep=True),
                'e_col': st.session_state['meta_e_color_df'].copy(deep=True)
            })
            if len(hist) > 4: hist.pop(0)
            st.session_state['schedule_history'] = hist

            st.session_state['generated_df'].at[chosen_sid, hole_date] = hole_shift
            st.session_state['meta_color_df'].at[chosen_sid, hole_date] = 'black'

            st.session_state['editor_reset_key'] += 1
            st.success(f"✅ {col_map_local[hole_date]} の {hole_shift} に、{chosen_name} さんを補充しました（下書き状態）")
            st.rerun()
        else:
            st.error("補充できるスタッフ（上限回数を超えず、休み希望も出していない人）が見つかりませんでした。手動での調整をお願いします。")
# ==========================================
# 3. メインアプリ UI構築
# ==========================================

st.title("🏥 コスモス外来・OPEシフトシステム")

staff_df = load_or_create_staff_data()
current_wishes = load_wishes_data()
current_remarks = load_remarks_data() 
current_committees = load_committees_data()
current_closed_days = load_closed_days_data() 

app_mode = st.sidebar.radio("モード選択", ["👤 スタッフ入力画面", "🔧 師長用メニュー"], index=0)

if app_mode != st.session_state['previous_app_mode']:
    st.session_state['admin_logged_in'] = False
    st.session_state['staff_logged_in_id'] = None
    st.session_state['previous_app_mode'] = app_mode
    st.rerun()

today = datetime.date.today()

# 15日を境目に、デフォルト表示する「対象月」を自動切り替え
if today.day <= DEADLINE_DAY:
    offset_months = 1  # 1日〜15日は「来月」を表示
else:
    offset_months = 2  # 16日〜月末は「再来月」を表示

target_month_raw = today.month + offset_months
default_year = today.year + (target_month_raw - 1) // 12
default_month = (target_month_raw - 1) % 12 + 1

st.sidebar.markdown("---")
st.sidebar.caption("対象年月設定")
selected_year = st.sidebar.number_input("年", min_value=2024, max_value=2030, value=default_year)
selected_month = st.sidebar.number_input("月", min_value=1, max_value=12, value=default_month)

dates = get_dates_in_month(selected_year, selected_month)
date_strs = [d.strftime('%Y-%m-%d') for d in dates]

jp_weekdays = ['月', '火', '水', '木', '金', '土', '日']
short_dates = []
for d in dates:
    wd = jp_weekdays[d.weekday()]
    label = f"{d.day}({wd})"
    if jpholiday.is_holiday(d) or d.weekday() == 6: label = "🔴" + label
    elif d.weekday() == 5: label = "🔵" + label
    short_dates.append(label)

col_map = dict(zip(date_strs, short_dates))
rev_col_map = dict(zip(short_dates, date_strs))

# ★ サイドバーへの「備考リスト」の引っ越し (師長メニューのみ表示)
if app_mode == "🔧 師長用メニュー" and st.session_state['admin_logged_in']:
    st.sidebar.markdown("---")
    st.sidebar.subheader("📝 今月の備考リスト")
    month_remarks = [r for r in current_remarks if r.get('year') == selected_year and r.get('month') == selected_month]
    if month_remarks:
        todo_remarks = [r for r in month_remarks if not r.get('done', False)]
        done_remarks = [r for r in month_remarks if r.get('done', False)]
        
        tab_todo, tab_done = st.sidebar.tabs([f"未反映 ({len(todo_remarks)})", f"反映済 ({len(done_remarks)})"])
        
        with tab_todo:
            if not todo_remarks:
                st.success("すべて処理済み！")
            for idx, r in enumerate(todo_remarks):
                with st.container(border=True):
                    is_done = st.checkbox("完了にする", value=False, key=f"sb_todo_{r['staff_id']}_{idx}")
                    if is_done:
                        for cr in current_remarks:
                            if cr['staff_id'] == r['staff_id'] and cr['year'] == r['year'] and cr['month'] == r['month'] and cr['remark'] == r['remark']:
                                cr['done'] = True
                        save_remarks_data(current_remarks)
                        st.rerun()
                    st.markdown(f"**{r['name']}**\n\n{r['remark']}")

        with tab_done:
            if not done_remarks:
                st.caption("反映済みの備考はありません。")
            for idx, r in enumerate(done_remarks):
                with st.container(border=True):
                    is_done = st.checkbox("✅ 完了", value=True, key=f"sb_done_{r['staff_id']}_{idx}")
                    if not is_done:
                        for cr in current_remarks:
                            if cr['staff_id'] == r['staff_id'] and cr['year'] == r['year'] and cr['month'] == r['month'] and cr['remark'] == r['remark']:
                                cr['done'] = False
                        save_remarks_data(current_remarks)
                        st.rerun()
                    safe_remark = r['remark'].replace('\n', '<br>')
                    st.markdown(f"<div style='text-decoration: line-through; color: #888888;'><b>{r['name']}</b><br>{safe_remark}</div>", unsafe_allow_html=True)
    else: 
        st.sidebar.caption("今月の備考はありません。")


if st.session_state.get('current_view_month') != f"{selected_year}-{selected_month:02d}":
    gen, col, e_gen, e_col = get_monthly_schedule(selected_year, selected_month)
    if gen is None:
        gen, col, e_gen, e_col = build_draft_schedule(selected_year, selected_month, staff_df, current_wishes, current_committees)
        save_monthly_schedule(selected_year, selected_month, gen, col, e_gen, e_col)

    st.session_state['generated_df'] = gen
    st.session_state['meta_color_df'] = col
    st.session_state['e_schedule_df'] = e_gen
    st.session_state['meta_e_color_df'] = e_col
    st.session_state['current_view_month'] = f"{selected_year}-{selected_month:02d}"
    st.session_state['schedule_history'] = []

# ---------------------------------------------------------------------
# 【A】 スタッフ入力画面
# ---------------------------------------------------------------------
if app_mode == "👤 スタッフ入力画面":
    st.header(f"📝 {selected_year}年{selected_month}月 希望調査")

    if selected_month == 1: 
        deadline_year, deadline_month = selected_year - 1, 12
    else: 
        deadline_year, deadline_month = selected_year, selected_month - 1
    
    deadline_date = datetime.date(deadline_year, deadline_month, DEADLINE_DAY)
    st.caption(f"提出期限: 毎月{DEADLINE_DAY}日 まで")

    if st.session_state['staff_logged_in_id'] is not None:
        c_logout, _ = st.columns([1, 6])
        with c_logout:
            if st.button("ログアウト", type="secondary"):
                st.session_state['staff_logged_in_id'] = None
                st.rerun()

    if st.session_state['staff_logged_in_id'] is None:
        st.info("職員番号とパスワードを入力してログインしてください。(初期パスワード: 1234)")
        with st.form("staff_login_form"):
            login_id_str = st.text_input("職員番号 (入力後Enter)", value="", placeholder="例: 1001")
            login_pass = st.text_input("パスワード", type="password")
            if st.form_submit_button("ログイン"):
                try:
                    login_id = int(login_id_str)
                    user_record = staff_df[staff_df["職員番号"] == login_id]
                    if not user_record.empty:
                        if login_pass == str(user_record.iloc[0]["パスワード"]):
                            st.session_state['staff_logged_in_id'] = login_id
                            st.success("ログイン成功")
                            st.rerun()
                        else: st.error("パスワードが違います")
                    else: st.error("職員番号が見つかりません")
                except ValueError: st.error("職員番号は数字で入力してください")
        st.stop()

    my_id = st.session_state['staff_logged_in_id']
    
    matching_staff = staff_df[staff_df["職員番号"].astype(str) == str(my_id)]
    if matching_staff.empty:
        st.session_state['staff_logged_in_id'] = None
        st.warning("職員データが見つかりません。名簿が更新された可能性があります。再度ログインしてください。")
        st.rerun()
        
    my_info = matching_staff.iloc[0]
    my_name, my_role, my_group = my_info["名前"], my_info["役職"], my_info.get("所属", "外来")
    st.success(f"ようこそ、**{my_name}** ({my_group} / {my_role}) さん")

    with st.expander("🔑 パスワード変更はこちら"):
        with st.form("change_pass_form"):
            new_p1 = st.text_input("新しいパスワード", type="password")
            new_p2 = st.text_input("新しいパスワード(確認)", type="password")
            if st.form_submit_button("変更する"):
                if new_p1 and new_p1 == new_p2:
                    staff_df.loc[staff_df["職員番号"] == my_id, "パスワード"] = new_p1
                    save_staff_data(staff_df)
                    st.success("パスワードを変更しました")
                else: st.error("パスワードが一致しないか、空です")

    st.markdown("---")

    st.subheader("📊 現在の希望状況（看護師）")
    st.caption("※副師長・看護師の赤希望・青希望を合わせた合計人数です。調整の参考にしてください。")
    ns_ids = staff_df[staff_df['役職'].isin(['副師長', '看護師'])]['職員番号'].tolist()
    all_month_wishes = [w for w in current_wishes if w['date'] in date_strs and w['staff_id'] in ns_ids]
    
    if all_month_wishes:
        wish_df_summary = pd.DataFrame(all_month_wishes)
        wish_counts = wish_df_summary.groupby(['date', 'shift']).size().unstack(fill_value=0)
        wish_counts = wish_counts.reindex(index=date_strs, columns=SHIFT_TYPES, fill_value=0)
        wish_counts_display = wish_counts.T
        wish_counts_display = wish_counts_display.loc[(wish_counts_display > 0).any(axis=1)]
        wish_counts_display.rename(columns=col_map, inplace=True)
        st.dataframe(wish_counts_display.replace(0, ""), use_container_width=True)
    else:
        st.info("現在、提出されている希望はありません。")

    st.markdown("---")

    is_past_deadline = today > deadline_date
    
    if is_past_deadline:
        st.error(f"⚠️ {selected_year}年{selected_month}月の勤務希望は締め切られました。（期限: {deadline_date.month}月{deadline_date.day}日）\n\n勤務希望の変更がある場合は師長まで報告してください")
        st.subheader("提出済みの希望")
        my_wishes = [w for w in current_wishes if w['staff_id'] == my_id]
        if my_wishes:
            my_wishes_df = pd.DataFrame(my_wishes)
            my_wishes_df = my_wishes_df[my_wishes_df['date'].isin(date_strs)]
            if not my_wishes_df.empty:
                my_wishes_df = my_wishes_df[['date', 'shift', 'type']].sort_values('date')
                my_wishes_df['type'] = my_wishes_df['type'].map({'red': '🔴 赤希望', 'blue': '🔵 青希望'})
                st.dataframe(my_wishes_df, use_container_width=True, hide_index=True)
            else: st.write(f"{selected_month}月の提出された希望はありません。")
        else: st.write("提出された希望はありません。")
            
        my_remark_current = next((r['remark'] for r in current_remarks if r['staff_id'] == my_id and r['year'] == selected_year and r['month'] == selected_month), "")
        if my_remark_current:
            st.write("**📝 登録済みの備考**")
            st.info(my_remark_current)
    else:
        st.subheader("希望入力フォーム")
        st.info("※保存されている希望があれば、自動的にフォームに入力されます。修正して「登録」を押してください。")
        st.caption("希望がない枠は「- (なし)」のままにしてください。\n※健診の希望がある場合は、希望枠ではなく下の「備考」に入力してください。")

        my_shift_options = ["-"] + get_allowed_shifts(my_role)
        my_shift_options = [s for s in my_shift_options if s not in ['OP日', '内日']]
        
        my_raw_wishes = [w for w in current_wishes if w['staff_id'] == my_id and w['date'] in date_strs]
        restored_reds, restored_blues = restore_staff_inputs(my_raw_wishes)
        
        while len(restored_reds) < 2: restored_reds.append({"start": None, "shifts": ["-", "-", "-"]})
        while len(restored_blues) < 2: restored_blues.append({"date": None, "shift": "-"})

        input_red_wishes = []
        input_blue_wishes = []

        st.markdown("#### 🔴 赤希望 (連続3日間 × 2セット)")
        for set_i in range(2):
            with st.container(border=True):
                c_label, c_clear = st.columns([4, 1])
                with c_label: st.write(f"**赤希望セット {set_i+1}**")
                with c_clear:
                    if st.button("クリア", key=f"clear_red_{set_i}"):
                        st.session_state[f"red_start_{set_i}"] = None
                        for k in range(3): st.session_state[f"red_shift_{set_i}_{k}"] = "-"
                        st.rerun()

                def_date = restored_reds[set_i]["start"]
                if def_date is not None and not (dates[0] <= def_date <= dates[-1]): def_date = None
                
                if f"red_start_{set_i}" in st.session_state and st.session_state[f"red_start_{set_i}"] is None:
                    def_date = None

                start_date = st.date_input(f"開始日", value=def_date, min_value=dates[0], max_value=dates[-1], key=f"red_start_{set_i}", label_visibility="collapsed")
                
                use_default_shifts = (start_date == def_date and def_date is not None)
                
                cols = st.columns(3)
                if start_date:
                    target_dates = [start_date + datetime.timedelta(days=i) for i in range(3)]
                    for i, d in enumerate(target_dates):
                        with cols[i]:
                            w_label = ['月', '火', '水', '木', '金', '土', '日'][d.weekday()]
                            d_label = f"{d.month}/{d.day}({w_label})"
                            disabled = False
                            if d not in dates:
                                st.caption(f"{d_label} (月外)")
                                disabled = True
                            else:
                                if is_holiday_jp(d) or d.weekday() == 6: st.markdown(f":red[{d_label}]")
                                elif d.weekday() == 5: st.markdown(f":blue[{d_label}]")
                                else: st.write(f"{d_label}")
                            
                            def_shift = restored_reds[set_i]["shifts"][i] if use_default_shifts else "-"
                            if def_shift in ['OP日', '内日']: def_shift = '日'
                            idx = 0
                            if def_shift in my_shift_options: idx = my_shift_options.index(def_shift)
                            
                            s = st.selectbox(f"シフト", my_shift_options, index=idx, key=f"red_shift_{set_i}_{i}", disabled=disabled, label_visibility="collapsed")
                            if not disabled and s != "-": input_red_wishes.append({"date": d, "shift": s, "type": "red"})
                else: st.caption("開始日を選択してください")

        st.markdown("#### 🔵 青希望 (単発 × 2回)")
        blue_cols = st.columns(2)
        for set_i in range(2):
            with blue_cols[set_i]:
                with st.container(border=True):
                    c_head, c_clr = st.columns([3, 1])
                    with c_head: st.write(f"**青希望 {set_i+1}**")
                    with c_clr:
                        if st.button("クリア", key=f"clear_blue_{set_i}"):
                            st.session_state[f"blue_d_{set_i}"] = None
                            st.session_state[f"blue_s_{set_i}"] = "-"
                            st.rerun()
                    
                    def_b_date = restored_blues[set_i]["date"]
                    def_b_shift = restored_blues[set_i]["shift"]
                    if def_b_date is not None and not (dates[0] <= def_b_date <= dates[-1]):
                        def_b_date = None
                        def_b_shift = "-"
                        
                    if f"blue_d_{set_i}" in st.session_state and st.session_state[f"blue_d_{set_i}"] is None:
                        def_b_date = None
                    
                    bd = st.date_input("日付", value=def_b_date, min_value=dates[0], max_value=dates[-1], key=f"blue_d_{set_i}")
                    
                    use_def_b = (bd == def_b_date and def_b_date is not None)
                    final_def_b_shift = def_b_shift if use_def_b else "-"
                    if final_def_b_shift in ['OP日', '内日']: final_def_b_shift = '日'
                    
                    idx_b = 0
                    if final_def_b_shift in my_shift_options: idx_b = my_shift_options.index(final_def_b_shift)
                    bs = st.selectbox("シフト", my_shift_options, index=idx_b, key=f"blue_s_{set_i}")
                    if bd and bs != "-": input_blue_wishes.append({"date": bd, "shift": bs, "type": "blue"})

        st.markdown("#### 💬 備考")
        my_remark = next((r['remark'] for r in current_remarks if r['staff_id'] == my_id and r['year'] == selected_year and r['month'] == selected_month), "")
        input_remark = st.text_area("委員会、健診、シフトに関する要望など", value=my_remark, height=100)

        st.markdown("---")
        if st.button("上記の内容で希望を登録（上書き）", type="primary"):
            all_inputs = input_red_wishes + input_blue_wishes
            errors, valid_records = [], []
            for item in all_inputs:
                shift_to_save = item['shift']
                if shift_to_save == '日':
                    if not is_holiday_jp(item['date']):
                        if my_group == 'オペ室': shift_to_save = 'OP日'
                        elif my_group == '内視鏡': shift_to_save = '内日'

                err = check_shift_rule(my_role, item['date'], shift_to_save)
                if err: errors.append(err)
                else: valid_records.append({'staff_id': int(my_id), 'name': my_name, 'date': item['date'].strftime('%Y-%m-%d'), 'shift': shift_to_save, 'type': item['type']})
            
            if errors:
                for e in errors: st.error(e)
                st.error("エラーがあるため登録できませんでした。")
            else:
                updated_wishes = [w for w in current_wishes if w['staff_id'] != my_id]
                updated_wishes.extend(valid_records)
                save_wish_data(updated_wishes)
                current_wishes = updated_wishes
                
                updated_remarks = [r for r in current_remarks if not (r['staff_id'] == my_id and r['year'] == selected_year and r['month'] == selected_month)]
                if input_remark.strip(): updated_remarks.append({'staff_id': int(my_id), 'name': my_name, 'year': int(selected_year), 'month': int(selected_month), 'remark': input_remark.strip(), 'done': False})
                save_remarks_data(updated_remarks)
                show_registration_success(valid_records, input_remark.strip())

    st.markdown("---")
    st.subheader("現在の登録状況")
    my_wishes = [w for w in current_wishes if w['staff_id'] == my_id]
    if my_wishes:
        my_wishes_df = pd.DataFrame(my_wishes)
        my_wishes_df = my_wishes_df[my_wishes_df['date'].isin(date_strs)]
        if not my_wishes_df.empty:
            my_wishes_df = my_wishes_df[['date', 'shift', 'type']].sort_values('date')
            my_wishes_df['type'] = my_wishes_df['type'].map({'red': '🔴 赤希望', 'blue': '🔵 青希望'})
            st.dataframe(my_wishes_df, use_container_width=True, hide_index=True) 
        else: st.write(f"{selected_month}月の登録希望はありません。")
    else: st.caption("現在登録されている希望はありません。")
        
    my_remark_current = next((r['remark'] for r in current_remarks if r['staff_id'] == my_id and r['year'] == selected_year and r['month'] == selected_month), "")
    if my_remark_current:
        st.write("**📝 登録済みの備考**")
        st.info(my_remark_current)


# ---------------------------------------------------------------------
# 【B】 管理者メニュー (師長用メニュー)
# ---------------------------------------------------------------------
elif app_mode == "🔧 師長用メニュー":
    if not st.session_state['admin_logged_in']:
        st.header("🔒 師長用 認証")
        with st.form("admin_login"):
            input_pass = st.text_input("パスワードを入力してください", type="password")
            if st.form_submit_button("ログイン"):
                if input_pass == ADMIN_PASSWORD:
                    st.session_state['admin_logged_in'] = True
                    
                    # 🌟【ここを追加！】ログイン成功と同時に編集ロックをONにする
                    st.session_state['schedule_locked'] = True
                    
                    st.success("ログインしました")
                    st.rerun()
                else: st.error("パスワードが違います")
        st.stop()
    
    col_logout, _ = st.columns([1, 5])
    with col_logout:
        if st.button("ログアウト (師長用)"):
            st.session_state['admin_logged_in'] = False
            st.rerun()

    st.header("🔧 師長用ダッシュボード")
    st.session_state['wishes'] = current_wishes

    tab_shift, tab_master = st.tabs(["📅 シフト作成・編集", "👤 スタッフマスタ管理"])

# === スタッフマスタ管理 ===
    with tab_master:
        current_staff_df = load_or_create_staff_data()
        mode = st.radio("操作", ["🟢 新規登録", "📝 一覧一括編集 (並び替え・修正)", "🔴 削除"], horizontal=True)
        limit_opts = ["無し"] + [str(i) for i in range(11)]
        
        if mode == "🟢 新規登録":
            with st.form("add_staff_form"):
                new_id = st.number_input("職員番号", min_value=1, step=1, value=None)
                new_name = st.text_input("名前")
                new_group = st.selectbox("所属", GROUPS) 
                new_role = st.selectbox("役職", ROLES)
                new_order = st.number_input("表示順 (小さいほど上に表示)", value=999)
                col_lim1, col_lim2, col_lim3, col_lim4 = st.columns(4)
                with col_lim1: new_night_limit = st.number_input("夜勤上限回数", min_value=0, value=4)
                with col_lim2: new_hol_day_limit = st.selectbox("休日日勤上限", limit_opts)
                with col_lim3: new_late_limit = st.selectbox("遅出半遅上限 (月間)", limit_opts)
                with col_lim4: new_e_limit = st.number_input("待機(E)上限", min_value=0, value=0)
                new_student = st.checkbox("🎓 学生 (日曜・祝日のみ公休計算)")
                st.caption("※初期パスワードは自動的に「1234」に設定されます。E上限が0の場合は待機担当に割り当てられません。")
                if st.form_submit_button("登録確認へ"):
                    if new_id and new_name:
                        if new_id in current_staff_df["職員番号"].values: st.error("番号重複")
                        else:
                            confirm_registration(new_id, new_name, new_group, new_role, new_order, new_night_limit, new_hol_day_limit, new_late_limit, new_e_limit, new_student, current_staff_df)
                    else: st.error("必須項目不足")

        elif mode == "📝 一覧一括編集 (並び替え・修正)":
            st.write("💡 表のセルを直接クリックして情報を書き換えられます。右端の専門スキルや学生設定もチェックボックスで付与できます。")
            edit_staff_df = current_staff_df.copy()
            edited_df = st.data_editor(
                edit_staff_df,
                column_config={
                    "職員番号": st.column_config.NumberColumn("職員番号", disabled=True),
                    "名前": st.column_config.TextColumn("名前"),
                    "所属": st.column_config.SelectboxColumn("所属", options=GROUPS), 
                    "役職": st.column_config.SelectboxColumn("役職", options=ROLES),
                    "表示順": st.column_config.NumberColumn("表示順"),
                    "夜勤上限": st.column_config.NumberColumn("夜勤上限", min_value=0),
                    "休日日勤上限": st.column_config.SelectboxColumn("休日日勤上限", options=limit_opts),
                    "遅出半遅上限": st.column_config.SelectboxColumn("遅出半遅上限", options=limit_opts),
                    "待機上限": st.column_config.NumberColumn("待機(E)上限", min_value=0),
                    "パスワード": st.column_config.TextColumn("パスワード"),
                    "DM": st.column_config.CheckboxColumn("DM"),
                    "リブレ": st.column_config.CheckboxColumn("リブレ"),
                    "学生": st.column_config.CheckboxColumn("学生(日祝公休)"),
                    # 🌟【追加】基本勤務を選択肢から選べるようにする
                    "基本勤務": st.column_config.SelectboxColumn("基本勤務", options=["日", "P4", "P5", "P8", "P11", "P13", "P17", "P70", "内P7", "内P9", "内P16"])
                },
                hide_index=True, use_container_width=True
            )
            if st.button("一括保存する", type="primary"):
                if not edited_df.equals(current_staff_df):
                    edited_df = edited_df.sort_values('表示順')
                    save_staff_data(edited_df)
                    st.success("スタッフ情報を更新し、並び替えを保存しました。")
                    st.rerun()
                else: st.info("変更されていません。")

        elif mode == "🔴 削除":
            staff_choices = {row["職員番号"]: f"{row['職員番号']}: {row['名前']}" for _, row in current_staff_df.iterrows()}
            target_id_del = st.selectbox("削除対象", sorted(staff_choices.keys()), format_func=lambda x: staff_choices[x])
            if st.button("削除確認へ", type="primary"): confirm_delete(target_id_del, staff_choices[target_id_del], current_staff_df)

    # === シフト作成・編集 ===
    with tab_shift:
        
        # 🌟 ロック機能の追加
        st.markdown("---")
        st.subheader("🔒 シフト編集ロック")
        is_locked = st.checkbox("編集ロックをかける（誤操作による書き換えを防ぐための閲覧専用モードになります）", value=st.session_state.get('schedule_locked', False))
        st.session_state['schedule_locked'] = is_locked
        if is_locked:
            st.success("🔒 現在、シフトはロックされています。編集や自動生成を行う場合はチェックを外してください。")
        
        # ★【折りたたみ追加】
        with st.expander("📅 行事・委員会・休診日の設定", expanded=False):
            col_c_form, col_c_list = st.columns([1, 1])
            
            with col_c_form:
                st.write("▼ 委員会の登録")
                with st.form("committee_form"):
                    c_date = st.date_input("日付", min_value=dates[0], max_value=dates[-1])
                    c_name_sel = st.selectbox("委員会名", ["代行", "記録", "教育", "その他"])
                    c_name_other = st.text_input("その他の場合")
                    
                    staff_dict_for_c = {row["職員番号"]: row["名前"] for _, row in staff_df.iterrows()}
                    c_mems = st.multiselect("メンバーを選択", options=list(staff_dict_for_c.keys()), format_func=lambda x: staff_dict_for_c[x])
                    
                    if st.form_submit_button("委員会を追加"):
                        final_c_name = c_name_other.strip() if c_name_sel == "その他" and c_name_other.strip() else c_name_sel
                        if not final_c_name or not c_mems:
                            st.error("委員会名とメンバーを入力してください。")
                        else:
                            new_c = {
                                "year": selected_year, "month": selected_month,
                                "date": c_date.strftime('%Y-%m-%d'), "committee": final_c_name, "members": c_mems
                            }
                            current_committees.append(new_c)
                            save_committees_data(current_committees)
                            st.success(f"{final_c_name} を追加しました！")
                            st.rerun()
                            
                st.write("▼ フリーメモ")
                my_event = next((e['note'] for e in load_events_data() if e['year'] == selected_year and e['month'] == selected_month), "")
                with st.form("event_form"):
                    event_input = st.text_area("委員会以外の予定・連絡事項", value=my_event, height=100, placeholder="例: 20日はワックス掛け")
                    if st.form_submit_button("メモを保存"):
                        updated_events = [e for e in load_events_data() if not (e['year'] == selected_year and e['month'] == selected_month)]
                        if event_input.strip(): updated_events.append({'year': selected_year, 'month': selected_month, 'note': event_input.strip()})
                        save_events_data(updated_events)
                        st.success("メモを保存しました！")
                        st.rerun()

            with col_c_list:
                st.write("▼ 登録済みの委員会")
                month_comms = [c for c in current_committees if c['year'] == selected_year and c['month'] == selected_month]
                if month_comms:
                    for idx, c in enumerate(month_comms):
                        with st.container(border=True):
                            c_col1, c_col2 = st.columns([4, 1])
                            d_str = datetime.datetime.strptime(c['date'], '%Y-%m-%d').strftime('%m/%d')
                            mem_names = [staff_dict_for_c.get(sid, "不明") for sid in c['members']]
                            c_col1.write(f"🟡 **{d_str} : {c['committee']}**")
                            c_col1.caption(f"参加: {', '.join(mem_names)}")
                            if c_col2.button("削除", key=f"del_comm_{idx}"):
                                current_committees.remove(c)
                                save_committees_data(current_committees)
                                st.rerun()
                else:
                    st.caption("登録されている委員会はありません。")

                st.write("▼ 専門外来の休診日")
                with st.form("closed_day_form"):
                    cd_col1, cd_col2, cd_col3 = st.columns([2, 2, 1])
                    with cd_col1: cd_date = st.date_input("休診日", min_value=dates[0], max_value=dates[-1], label_visibility="collapsed")
                    with cd_col2: cd_type = st.selectbox("休診する専門外来", ["DM", "リブレ"], label_visibility="collapsed")
                    with cd_col3:
                        if st.form_submit_button("登録"):
                            new_cd = {
                                "year": selected_year, "month": selected_month,
                                "date": cd_date.strftime('%Y-%m-%d'), "clinic_type": cd_type
                            }
                            if not any(c['date'] == new_cd['date'] and c['clinic_type'] == new_cd['clinic_type'] for c in current_closed_days):
                                current_closed_days.append(new_cd)
                                save_closed_days_data(current_closed_days)
                                st.rerun()
                            else:
                                st.error("登録済み")

                month_closed = [c for c in current_closed_days if c['year'] == selected_year and c['month'] == selected_month]
                if month_closed:
                    for idx, c in enumerate(month_closed):
                        with st.container(border=True):
                            c_col1, c_col2 = st.columns([4, 1])
                            d_str = datetime.datetime.strptime(c['date'], '%Y-%m-%d').strftime('%m/%d')
                            c_col1.write(f"🚫 **{d_str} : {c['clinic_type']} 外来休診**")
                            if c_col2.button("削除", key=f"del_cd_{idx}"):
                                current_closed_days.remove(c)
                                save_closed_days_data(current_closed_days)
                                st.rerun()
                else:
                    st.caption("登録されている休診日はありません。")

        # ★【折りたたみ追加】
        with st.expander("⚙️ 外来 日勤必要人数設定 (Ns等)", expanded=False):
            st.caption("※日勤計は休日は「師長」も含めて計算されます。土曜は遅出1名、日曜・祝日は半/遅1名を絶対条件とします。")
            default_counts = []
            for d in dates:
                is_sun_or_hol = d.weekday() == 6 or jpholiday.is_holiday(d)
                is_sat_only = d.weekday() == 5 and not jpholiday.is_holiday(d)
                
                if is_sun_or_hol: 
                    req_am, req_pm = 3, 3
                    req_late, req_hl = 0, 1
                elif is_sat_only: 
                    req_am, req_pm = 2, 2
                    req_late, req_hl = 1, 0
                else: 
                    req_am, req_pm = 10, 10
                    req_late, req_hl = 0, 0
                    
                day_label = f"{d.day} ({['月','火','水','木','金','土','日'][d.weekday()]})"
                default_counts.append({"日付": d, "曜日": day_label, "Ns必要数(AM)": req_am, "Ns必要数(PM)": req_pm, "遅出必要数": req_late, "半/遅必要数": req_hl})
                
            req_df = pd.DataFrame(default_counts)
            edited_req_df = st.data_editor(req_df[["曜日", "Ns必要数(AM)", "Ns必要数(PM)", "遅出必要数", "半/遅必要数"]], use_container_width=True, disabled=["曜日"], hide_index=True)
            req_dict = {req_df.iloc[i]["日付"].strftime('%Y-%m-%d'): {"am": row["Ns必要数(AM)"], "pm": row["Ns必要数(PM)"], "late": row["遅出必要数"], "half_late": row["半/遅必要数"]} for i, row in edited_req_df.iterrows()}

        # ★【折りたたみ追加】
        with st.expander("📋 集約された希望リスト", expanded=False):
            if current_wishes:
                admin_wish_df = pd.DataFrame(current_wishes)
                admin_wish_df = admin_wish_df[admin_wish_df['date'].isin(date_strs)]
                if not admin_wish_df.empty:
                    admin_wish_df = admin_wish_df.sort_values(by=['date', 'staff_id'])
                    st.dataframe(admin_wish_df, use_container_width=True, hide_index=True)
                else: st.info(f"{selected_month}月の希望はまだ提出されていません")
            else: st.info("希望はまだ提出されていません")

        holiday_day_counts = {sid: 0 for sid in staff_df["職員番号"].values}

        # ==========================================
        # ★ 生成・操作パネル
        # ==========================================
        if not is_locked:
            st.markdown("---")
            st.subheader("⚙️ シフト自動生成")
            
            col_gen1, col_gen2 = st.columns(2)
            with col_gen1:
                if st.button("🔄 空き枠を自動生成 (緑ロックを保持)", type="primary", use_container_width=True):
                    confirm_generation_dialog("keep")
            with col_gen2:
                if st.button("⚠️ シフトを初期化 (下書きに戻す)", use_container_width=True):
                    confirm_generation_dialog("reset")
            if st.session_state.get('run_generation'):
                gen_mode = st.session_state['run_generation']
                st.session_state['run_generation'] = None 
                
                gen_keep = (gen_mode == "keep")
                gen_reset = (gen_mode == "reset")

                with st.spinner("処理中..."):
                    if st.session_state.get('generated_df') is not None:
                        hist = st.session_state['schedule_history']
                        hist.append({
                            'gen': st.session_state['generated_df'].copy(deep=True),
                            'col': st.session_state['meta_color_df'].copy(deep=True),
                            'e_gen': st.session_state['e_schedule_df'].copy(deep=True),
                            'e_col': st.session_state['meta_e_color_df'].copy(deep=True)
                        })
                        if len(hist) > 4: hist.pop(0)
                        st.session_state['schedule_history'] = hist
                    
                    if gen_reset:
                        gen, col, e_gen, e_col = build_draft_schedule(selected_year, selected_month, staff_df, current_wishes, current_committees)
                        st.session_state['generated_df'] = gen
                        st.session_state['meta_color_df'] = col
                        st.session_state['e_schedule_df'] = e_gen
                        st.session_state['meta_e_color_df'] = e_col
                        persist_current_schedule(selected_year, selected_month)
                        st.success("✅ 下書き状態にリセットしました！")
                        st.rerun()

                    cols = [d.strftime('%Y-%m-%d') for d in dates]
                    sids = staff_df["職員番号"].values
                    schedule = pd.DataFrame(index=sids, columns=cols)
                    meta_color = pd.DataFrame("black", index=sids, columns=cols)
                    staff_info = staff_df.set_index("職員番号").to_dict(orient="index")
                    
                    apply_rollover(schedule, meta_color, selected_year, selected_month, sids, cols)
                    
                    all_ns_ids = [k for k,v in staff_info.items() if v['役職'] in ['副師長','看護師']]
                    tgt_ids_holiday = [k for k,v in staff_info.items() if v['役職'] in ['師長', '副師長','看護師']]
                    
                    e_schedule = pd.DataFrame(index=['E'], columns=cols)
                    meta_e_color = pd.DataFrame("black", index=['E'], columns=cols)
                    
                    night_counts = {sid: 0 for sid in sids}
                    work_counts = {sid: 0 for sid in sids}
                    half_late_counts = {sid: 0 for sid in sids}
                    late_counts = {sid: 0 for sid in sids}
                    am_shifts = ['日', '/年', '/半', '/特', '/夏', '/研', 'P4', 'P5', 'P8', 'P11', 'P13', 'P17', 'P70']
                    
                    num_holidays = sum(1 for d in dates if is_holiday_jp(d))
                    sun_hol_count = sum(1 for d in dates if d.weekday() == 6 or jpholiday.is_holiday(d))
                    target_offs_dict = {sid: sun_hol_count if staff_info[sid].get('学生', False) else num_holidays for sid in sids}
                    
                    holiday_day_counts = {sid: 0 for sid in sids}

                    if gen_keep and st.session_state.get('generated_df') is not None and st.session_state.get('meta_color_df') is not None:
                        old_schedule = st.session_state['generated_df']
                        old_color = st.session_state['meta_color_df']
                        for s in sids:
                            if s in old_schedule.index and s in old_color.index:
                                for c in cols:
                                    if c in old_schedule.columns and c in old_color.columns:
                                        if old_color.at[s, c] == 'green':
                                            val = old_schedule.at[s, c]
                                            schedule.at[s, c] = val
                                            meta_color.at[s, c] = 'green'
                                            
                                            if val == 'A':
                                                cd = datetime.datetime.strptime(c, '%Y-%m-%d').date()
                                                if cd in dates:
                                                    idx = dates.index(cd)
                                                    if idx + 1 < len(dates):
                                                        nd_s = dates[idx+1].strftime('%Y-%m-%d')
                                                        if pd.isna(schedule.at[s, nd_s]) or old_color.at[s, nd_s] != 'green': schedule.at[s, nd_s] = 'B'
                                                    if idx + 2 < len(dates):
                                                        nd2_s = dates[idx+2].strftime('%Y-%m-%d')
                                                        if pd.isna(schedule.at[s, nd2_s]) or old_color.at[s, nd2_s] != 'green': schedule.at[s, nd2_s] = 'OFF'
                                            elif val == 'B':
                                                cd = datetime.datetime.strptime(c, '%Y-%m-%d').date()
                                                if cd in dates:
                                                    idx = dates.index(cd)
                                                    if idx - 1 >= 0:
                                                        pd_s = dates[idx-1].strftime('%Y-%m-%d')
                                                        if pd.isna(schedule.at[s, pd_s]) or old_color.at[s, pd_s] != 'green': schedule.at[s, pd_s] = 'A'
                                                    if idx + 1 < len(dates):
                                                        nd_s = dates[idx+1].strftime('%Y-%m-%d')
                                                        if pd.isna(schedule.at[s, nd_s]) or old_color.at[s, nd_s] != 'green': schedule.at[s, nd_s] = 'OFF'
                                            elif val == '遅':
                                                cd = datetime.datetime.strptime(c, '%Y-%m-%d').date()
                                                late_counts[s] += 1
                                                if cd.weekday() == 5 and not jpholiday.is_holiday(cd):
                                                    if cd in dates:
                                                        idx = dates.index(cd)
                                                        if idx + 1 < len(dates):
                                                            nd_s = dates[idx+1].strftime('%Y-%m-%d')
                                                            if pd.isna(schedule.at[s, nd_s]) and meta_color.at[s, nd_s] != 'green':
                                                                if schedule.at[s, nd_s] not in REST_AND_A_SHIFTS: schedule.at[s, nd_s] = 'OFF'
                                            elif val in ['半/遅', '年/遅']:
                                                cd = datetime.datetime.strptime(c, '%Y-%m-%d').date()
                                                half_late_counts[s] += 1
                                                if cd.weekday() == 6 or jpholiday.is_holiday(cd):
                                                    if cd in dates:
                                                        idx = dates.index(cd)
                                                        if idx + 1 < len(dates):
                                                            nd_s = dates[idx+1].strftime('%Y-%m-%d')
                                                            if pd.isna(schedule.at[s, nd_s]) and meta_color.at[s, nd_s] != 'green':
                                                                if schedule.at[s, nd_s] not in REST_AND_A_SHIFTS: schedule.at[s, nd_s] = 'OFF'
                                                                
                                            if val == '日' and is_holiday_jp(datetime.datetime.strptime(c, '%Y-%m-%d').date()): holiday_day_counts[s] += 1
                                                
                        old_e_schedule = st.session_state.get('e_schedule_df')
                        old_e_color = st.session_state.get('meta_e_color_df')
                        if old_e_schedule is not None and old_e_color is not None:
                            for c in cols:
                                if old_e_color.at['E', c] == 'green':
                                    e_schedule.at['E', c] = old_e_schedule.at['E', c]
                                    meta_e_color.at['E', c] = 'green'

                    committee_logs = []
                    for c in current_committees:
                        if c['year'] == selected_year and c['month'] == selected_month:
                            d_s = c['date']
                            if d_s in cols:
                                for sid in c['members']:
                                    if sid in schedule.index:
                                        if meta_color.at[sid, d_s] != 'green':
                                            wished_off = any(w['staff_id'] == sid and w['date'] == d_s and w['shift'] in ['OFF', '年休', '特休', '夏休', '/年', '年/', '/半', '半/', '年/半', '/特', '特/', '/夏', '夏/', '介護休暇', '病欠', '産休', '育休', '代休', '/代休'] for w in current_wishes)
                                            if not wished_off:
                                                if pd.isna(schedule.at[sid, d_s]):
                                                    # 🌟【修正】パートさんの基本勤務を優先する
                                                    b_s = staff_info[sid].get('基本勤務', '日')
                                                    if pd.isna(b_s) or b_s == '': b_s = '日'
                                                    
                                                    if b_s != '日':
                                                        def_c_shift = b_s
                                                    else:
                                                        def_c_shift = '日'
                                                        if staff_info[sid]['所属'] == 'オペ室': def_c_shift = 'OP日'
                                                        elif staff_info[sid]['所属'] == '内視鏡': def_c_shift = '内日'
                                                        
                                                    schedule.at[sid, d_s] = def_c_shift
                                                    staff_name = staff_info[sid]['名前']
                                                    short_d = datetime.datetime.strptime(d_s, '%Y-%m-%d').strftime('%m/%d')
                                                    committee_logs.append(f"・{short_d} : **{staff_name}**さん ({c['committee']})")

                    for w in current_wishes:
                        if w['type'] == 'red' and w['staff_id'] in schedule.index:
                            d_s = w['date']
                            if d_s not in cols: continue 
                            if meta_color.at[w['staff_id'], d_s] != 'green':
                                schedule.at[w['staff_id'], d_s] = w['shift']
                                meta_color.at[w['staff_id'], d_s] = 'red'
                                if w['shift'] == 'A':
                                    cd = datetime.datetime.strptime(d_s, '%Y-%m-%d').date()
                                    if cd in dates:
                                        idx = dates.index(cd)
                                        if idx + 1 < len(dates):
                                            nd_s = dates[idx+1].strftime('%Y-%m-%d')
                                            if pd.isna(schedule.at[w['staff_id'], nd_s]) and meta_color.at[w['staff_id'], nd_s] != 'green': schedule.at[w['staff_id'], nd_s] = 'B'
                                        if idx + 2 < len(dates):
                                            nd2_s = dates[idx+2].strftime('%Y-%m-%d')
                                            if pd.isna(schedule.at[w['staff_id'], nd2_s]) and meta_color.at[w['staff_id'], nd2_s] != 'green': schedule.at[w['staff_id'], nd2_s] = 'OFF'
                                elif w['shift'] == 'B':
                                    cd = datetime.datetime.strptime(d_s, '%Y-%m-%d').date()
                                    if cd in dates:
                                        idx = dates.index(cd)
                                        if idx - 1 >= 0:
                                            pd_s = dates[idx-1].strftime('%Y-%m-%d')
                                            if pd.isna(schedule.at[w['staff_id'], pd_s]) and meta_color.at[w['staff_id'], pd_s] != 'green': schedule.at[w['staff_id'], pd_s] = 'A'
                                        if idx + 1 < len(dates):
                                            nd_s = dates[idx+1].strftime('%Y-%m-%d')
                                            if pd.isna(schedule.at[w['staff_id'], nd_s]) and meta_color.at[w['staff_id'], nd_s] != 'green': schedule.at[w['staff_id'], nd_s] = 'OFF'
                                elif w['shift'] == '遅':
                                    late_counts[w['staff_id']] += 1
                                    cd = datetime.datetime.strptime(d_s, '%Y-%m-%d').date()
                                    if cd.weekday() == 5 and not jpholiday.is_holiday(cd):
                                        if cd in dates:
                                            idx = dates.index(cd)
                                            if idx + 1 < len(dates):
                                                nd_s = dates[idx+1].strftime('%Y-%m-%d')
                                                if pd.isna(schedule.at[w['staff_id'], nd_s]) and meta_color.at[w['staff_id'], nd_s] != 'green':
                                                    if schedule.at[w['staff_id'], nd_s] not in REST_AND_A_SHIFTS: schedule.at[w['staff_id'], nd_s] = 'OFF'
                                elif w['shift'] in ['半/遅', '年/遅']:
                                    half_late_counts[w['staff_id']] += 1
                                    cd = datetime.datetime.strptime(d_s, '%Y-%m-%d').date()
                                    if cd.weekday() == 6 or jpholiday.is_holiday(cd):
                                        if cd in dates:
                                            idx = dates.index(cd)
                                            if idx + 1 < len(dates):
                                                nd_s = dates[idx+1].strftime('%Y-%m-%d')
                                                if pd.isna(schedule.at[w['staff_id'], nd_s]) and meta_color.at[w['staff_id'], nd_s] != 'green':
                                                    if schedule.at[w['staff_id'], nd_s] not in REST_AND_A_SHIFTS: schedule.at[w['staff_id'], nd_s] = 'OFF'
                                
                                if w['shift'] == '日' and is_holiday_jp(datetime.datetime.strptime(d_s, '%Y-%m-%d').date()): holiday_day_counts[w['staff_id']] += 1

                    for s, inf in staff_info.items():
                        if inf['役職'] == '事務' or str(inf.get('休日日勤上限', '無し')) == '0':
                            for d in dates:
                                if is_holiday_jp(d):
                                    d_s = d.strftime('%Y-%m-%d')
                                    if pd.isna(schedule.at[s, d_s]): schedule.at[s, d_s] = 'OFF'

                    blue_w = [w for w in current_wishes if w['type'] == 'blue']
                    random.shuffle(blue_w)
                    for w in blue_w:
                        sid = w['staff_id']
                        if sid not in schedule.index: continue
                        d_str = w['date']
                        if d_str not in cols: continue
                        shift = w['shift']
                        role = staff_info[sid]['役職']
                        
                        if pd.notna(schedule.at[sid, d_str]): continue
                        valid = True
                        if shift == 'A' and role in ['師長', 'CW', '事務', 'CE']: valid = False
                        cur_obj = datetime.datetime.strptime(d_str, '%Y-%m-%d').date()
                        if role == '事務' and is_holiday_jp(cur_obj) and shift not in ['OFF', '年休', '特休', '夏休', '/年', '年/', '/半', '半/', '年/半', '/特', '特/', '/夏', '夏/', '介護休暇', '病欠', '産休', '育休', '代休', '/代休']: valid = False
                        
                        if shift == 'A':
                            if (schedule[d_str] == 'A').any(): valid = False
                            d_idx = dates.index(cur_obj)
                            if d_idx + 1 < len(dates):
                                n_str = dates[d_idx+1].strftime('%Y-%m-%d')
                                if pd.notna(schedule.at[sid, n_str]): valid = False
                        if shift == 'B':
                            d_idx = dates.index(cur_obj)
                            if d_idx - 1 >= 0:
                                p_str = dates[d_idx-1].strftime('%Y-%m-%d')
                                if pd.notna(schedule.at[sid, p_str]): valid = False
                            else: valid = False 
                                
                        if is_holiday_jp(cur_obj) and shift == '日':
                            if not can_work_holiday_day(sid, holiday_day_counts[sid], staff_info): valid = False
                        if shift in ['遅', '半/遅', '年/遅']:
                            total_late = late_counts[sid] + half_late_counts[sid]
                            if not can_work_late_shift(sid, total_late, staff_info): valid = False
                        
                        if valid:
                            schedule.at[sid, d_str] = shift
                            meta_color.at[sid, d_str] = 'blue'
                            if shift == '日' and is_holiday_jp(cur_obj): holiday_day_counts[sid] += 1
                            if shift == '遅': late_counts[sid] += 1
                            if shift in ['半/遅', '年/遅']: half_late_counts[sid] += 1
                            
                            if shift == 'A':
                                d_idx = dates.index(cur_obj)
                                if d_idx + 1 < len(dates):
                                    nd_s = dates[d_idx+1].strftime('%Y-%m-%d')
                                    if pd.isna(schedule.at[sid, nd_s]) and meta_color.at[sid, nd_s] != 'green': schedule.at[sid, nd_s] = 'B'
                                if d_idx + 2 < len(dates):
                                    nd2 = dates[d_idx+2].strftime('%Y-%m-%d')
                                    if pd.isna(schedule.at[sid, nd2]) and meta_color.at[sid, nd2] != 'green': schedule.at[sid, nd2] = 'OFF'
                            elif shift == 'B':
                                d_idx = dates.index(cur_obj)
                                if d_idx - 1 >= 0:
                                    pd_s = dates[d_idx-1].strftime('%Y-%m-%d')
                                    if pd.isna(schedule.at[sid, pd_s]) and meta_color.at[sid, pd_s] != 'green': schedule.at[sid, pd_s] = 'A'
                                if d_idx + 1 < len(dates):
                                    nd_s = dates[d_idx+1].strftime('%Y-%m-%d')
                                    if pd.isna(schedule.at[sid, nd_s]) and meta_color.at[sid, nd_s] != 'green': schedule.at[sid, nd_s] = 'OFF'
                            elif shift == '遅':
                                d_idx = dates.index(cur_obj)
                                if cur_obj.weekday() == 5 and not jpholiday.is_holiday(cur_obj):
                                    if d_idx + 1 < len(dates):
                                        nd_s = dates[d_idx+1].strftime('%Y-%m-%d')
                                        if pd.isna(schedule.at[sid, nd_s]) and meta_color.at[sid, nd_s] != 'green':
                                            if schedule.at[sid, nd_s] not in REST_AND_A_SHIFTS: schedule.at[sid, nd_s] = 'OFF'
                            elif shift in ['半/遅', '年/遅']:
                                d_idx = dates.index(cur_obj)
                                if cur_obj.weekday() == 6 or jpholiday.is_holiday(cur_obj):
                                    if d_idx + 1 < len(dates):
                                        nd_s = dates[d_idx+1].strftime('%Y-%m-%d')
                                        if pd.isna(schedule.at[sid, nd_s]) and meta_color.at[sid, nd_s] != 'green':
                                            if schedule.at[sid, nd_s] not in REST_AND_A_SHIFTS: schedule.at[sid, nd_s] = 'OFF'

                    libre_cws = [k for k,v in staff_info.items() if v.get('役職') == 'CW' and v.get('リブレ', False)]
                    dm_ns = [k for k,v in staff_info.items() if v.get('役職') in ['師長', '副師長', '看護師'] and v.get('DM', False)]
                    
                    for i, d in enumerate(dates):
                        d_s = d.strftime('%Y-%m-%d')
                        wd = d.weekday()
                        
                        closed_types_today = [c['clinic_type'] for c in current_closed_days if c['date'] == d_s]
                        
                        if not is_holiday_jp(d):
                            if wd in [1, 3]: 
                                if 'リブレ' not in closed_types_today:
                                    if libre_cws and not any(schedule.at[cw, d_s] in am_shifts for cw in libre_cws):
                                        cands = [cw for cw in libre_cws if pd.isna(schedule.at[cw, d_s])]
                                        if cands:
                                            random.shuffle(cands)
                                            schedule.at[cands[0], d_s] = '日'
                                            work_counts[cands[0]] += 1
                                            
                                if 'DM' not in closed_types_today:
                                    if dm_ns and not any(schedule.at[ns, d_s] in am_shifts for ns in dm_ns):
                                        cands = [ns for ns in dm_ns if pd.isna(schedule.at[ns, d_s])]
                                        if cands:
                                            random.shuffle(cands)
                                            schedule.at[cands[0], d_s] = '日'
                                            work_counts[cands[0]] += 1

                    def get_total_fixed_offs(target_sid):
                        cnt = 0
                        for wd in dates:
                            wd_s = wd.strftime('%Y-%m-%d')
                            val = schedule.at[target_sid, wd_s]
                            if pd.notna(val):
                                # ★ 代休・/代休はOFF計算に含めない
                                if val in ['OFF', '年休', '特休', '夏休', '介護休暇', '病欠', '産休', '育休']: cnt += 1
                                elif val in ['/半', '半/', '半/遅', '年/半', '健/半', '/夏', '夏/']: cnt += 0.5
                            else:
                                wished = [w for w in current_wishes if w['staff_id'] == target_sid and w['date'] == wd_s]
                                if wished:
                                    w_val = wished[0]['shift']
                                    if w_val in ['OFF', '年休', '特休', '夏休', '介護休暇', '病欠', '産休', '育休']: cnt += 1
                                    elif w_val in ['/半', '半/', '半/遅', '年/半', '健/半', '/夏', '夏/']: cnt += 0.5
                        return cnt

                    for sid in sids:
                        for col_d in cols:
                            val = schedule.at[sid, col_d]
                            if pd.notna(val):
                                if val in ['A', '当直']: night_counts[sid] += 1 
                                if val in WORK_SHIFTS: work_counts[sid] += 1

                    def get_night_dist(target_sid, current_idx):
                        dists = []
                        for j, ds in enumerate(cols):
                            if schedule.at[target_sid, ds] in ['A', '当直']:
                                dists.append(abs(current_idx - j))
                        return min(dists) if dists else 999

                    ncands = [k for k,v in staff_info.items() if v['役職'] in ['副師長','看護師']]
                    for i, d in enumerate(dates):
                        d_s = d.strftime('%Y-%m-%d')
                        if (schedule[d_s] == 'A').any(): continue
                        cands = []
                        for s in ncands:
                            night_limit = staff_info[s].get('夜勤上限', 4)
                            if pd.isna(schedule.at[s, d_s]) and night_counts[s] < night_limit:
                                if is_holiday_jp(d) and not can_work_holiday_day(s, holiday_day_counts[s], staff_info): continue
                                ok_next = True
                                if i+1 < len(dates):
                                    nd_s = dates[i+1].strftime('%Y-%m-%d')
                                    if pd.notna(schedule.at[s, nd_s]): ok_next = False
                                if i+2 < len(dates):
                                    nd2_s = dates[i+2].strftime('%Y-%m-%d')
                                    val = schedule.at[s, nd2_s]
                                    if pd.notna(val) and val not in ['OFF', '年休', '特休', '夏休', '介護休暇', '病欠', '産休', '育休']: ok_next = False
                                if ok_next: cands.append(s)
                                
                        if is_holiday_jp(d) and cands:
                            cands_limit_ok = [c for c in cands if can_work_holiday_day(c, holiday_day_counts[c], staff_info)]
                            safe_cands = [c for c in cands_limit_ok if not is_consecutive_holiday_work(schedule, c, i, dates)]
                            fallback_cands = [c for c in cands_limit_ok if is_consecutive_holiday_work(schedule, c, i, dates)]
                            
                            random.shuffle(safe_cands)
                            safe_cands.sort(key=lambda x: (night_counts[x], -get_night_dist(x, i)))
                            random.shuffle(fallback_cands)
                            fallback_cands.sort(key=lambda x: (night_counts[x], -get_night_dist(x, i)))
                            
                            cands_to_use = safe_cands + fallback_cands
                        else:
                            cands_to_use = cands
                            if cands_to_use:
                                random.shuffle(cands_to_use)
                                cands_to_use.sort(key=lambda x: (night_counts[x], -get_night_dist(x, i)))
                            
                        if cands_to_use:
                            chosen = cands_to_use[0]
                            schedule.at[chosen, d_s] = 'A'
                            night_counts[chosen] += 1 
                            work_counts[chosen] += 1
                            
                            if i+1 < len(dates): 
                                schedule.at[chosen, dates[i+1].strftime('%Y-%m-%d')] = 'B'
                                work_counts[chosen] += 1
                            if i+2 < len(dates): 
                                schedule.at[chosen, dates[i+2].strftime('%Y-%m-%d')] = 'OFF'

                    cw_ids = [k for k,v in staff_info.items() if v['役職']=='CW']
                    for i, d in enumerate(dates):
                        d_s = d.strftime('%Y-%m-%d')
                        if is_holiday_jp(d):
                            if (schedule.loc[cw_ids, d_s] == '日').sum() == 0:
                                cands = [s for s in cw_ids if pd.isna(schedule.at[s, d_s])]
                                cands_limit_ok = [c for c in cands if can_work_holiday_day(c, holiday_day_counts[c], staff_info)]
                                
                                safe_cands = [c for c in cands_limit_ok if not is_consecutive_holiday_work(schedule, c, i, dates)]
                                fallback_cands = [c for c in cands_limit_ok if is_consecutive_holiday_work(schedule, c, i, dates)]
                                
                                random.shuffle(safe_cands)
                                safe_cands.sort(key=lambda x: (1 if get_total_fixed_offs(x) >= target_offs_dict[x] else 0, work_counts[x]))
                                random.shuffle(fallback_cands)
                                fallback_cands.sort(key=lambda x: (1 if get_total_fixed_offs(x) >= target_offs_dict[x] else 0, work_counts[x]))
                                
                                cands_to_use = safe_cands + fallback_cands
                                
                                if cands_to_use:
                                    ch = cands_to_use[0]
                                    schedule.at[ch, d_s] = '日'
                                    holiday_day_counts[ch] += 1
                                    work_counts[ch] += 1

                    for s, inf in staff_info.items():
                        if inf['所属'] in ['オペ室', '内視鏡']:
                            # 🌟【追加】パートさんの基本勤務を取得
                            b_s = inf.get('基本勤務', '日')
                            if pd.isna(b_s) or b_s == '': b_s = '日'
                            
                            for i, d in enumerate(dates):
                                d_s = d.strftime('%Y-%m-%d')
                                if not is_holiday_jp(d) and pd.isna(schedule.at[s, d_s]):
                                    # パートさんなら「内P9」などを入れ、正職員なら「内日」「OP日」を入れる
                                    if b_s != '日':
                                        schedule.at[s, d_s] = b_s
                                    else:
                                        schedule.at[s, d_s] = 'OP日' if inf['所属'] == 'オペ室' else '内日'
                                    work_counts[s] += 1

                    late_hl_tgt_ids = [k for k,v in staff_info.items() if v['所属'] == '外来' and v['役職'] in ['副師長','看護師']]
                    
                    for i, d in enumerate(dates):
                        d_s = d.strftime('%Y-%m-%d')
                        req_late = req_dict[d_s]["late"]
                        req_hl = req_dict[d_s]["half_late"]
                        
                        is_sun_or_hol = d.weekday() == 6 or jpholiday.is_holiday(d)
                        is_sat_only = d.weekday() == 5 and not jpholiday.is_holiday(d)
                        
                        if is_holiday_jp(d):
                            current_late = (schedule.loc[tgt_ids_holiday, d_s] == '遅').sum() 
                        else:
                            current_late = (schedule.loc[all_ns_ids, d_s] == '遅').sum() 
                            
                        if current_late < req_late:
                            cands = [s for s in late_hl_tgt_ids if pd.isna(schedule.at[s, d_s])]
                            
                            if is_sat_only:
                                cands_limit_ok = [s for s in cands if can_work_late_shift(s, late_counts[s] + half_late_counts[s], staff_info)]
                                if i + 1 < len(dates):
                                    nd_s = dates[i+1].strftime('%Y-%m-%d')
                                    safe_cands_next = [s for s in cands_limit_ok if schedule.at[s, nd_s] in REST_AND_A_SHIFTS or pd.isna(schedule.at[s, nd_s])]
                                    if safe_cands_next: cands_limit_ok = safe_cands_next
                                
                                cands_to_use = cands_limit_ok.copy()
                                random.shuffle(cands_to_use)
                                cands_to_use.sort(key=lambda x: (
                                    1 if get_total_fixed_offs(x) >= target_offs_dict[x] else 0, 
                                    1 if is_consecutive_holiday_work(schedule, x, i, dates) else 0,
                                    get_total_fixed_offs(x), 
                                    late_counts[x]
                                ))
                            else: 
                                cands_to_use = []
                                
                            for k in range(min(req_late - current_late, len(cands_to_use))):
                                sid = cands_to_use[k]
                                schedule.at[sid, d_s] = '遅'
                                work_counts[sid] += 1
                                late_counts[sid] += 1
                                if is_sat_only and i + 1 < len(dates):
                                    nd_s = dates[i+1].strftime('%Y-%m-%d')
                                    if pd.isna(schedule.at[sid, nd_s]): schedule.at[sid, nd_s] = 'OFF'

                        if is_holiday_jp(d):
                            current_hl = (schedule.loc[tgt_ids_holiday, d_s].isin(['半/遅', '年/遅'])).sum() 
                        else:
                            current_hl = (schedule.loc[all_ns_ids, d_s].isin(['半/遅', '年/遅'])).sum() 
                            
                        if current_hl < req_hl:
                            cands = [s for s in late_hl_tgt_ids if pd.isna(schedule.at[s, d_s])]
                            
                            if is_sun_or_hol:
                                cands_limit_ok = [s for s in cands if can_work_late_shift(s, late_counts[s] + half_late_counts[s], staff_info)]
                                if i + 1 < len(dates):
                                    nd_s = dates[i+1].strftime('%Y-%m-%d')
                                    safe_cands_next = [s for s in cands_limit_ok if schedule.at[s, nd_s] in REST_AND_A_SHIFTS or pd.isna(schedule.at[s, nd_s])]
                                    if safe_cands_next: cands_limit_ok = safe_cands_next
                                    
                                cands_to_use = cands_limit_ok.copy()
                                random.shuffle(cands_to_use)
                                cands_to_use.sort(key=lambda x: (
                                    1 if get_total_fixed_offs(x) >= target_offs_dict[x] else 0, 
                                    1 if is_consecutive_holiday_work(schedule, x, i, dates) else 0,
                                    get_total_fixed_offs(x), 
                                    half_late_counts[x]
                                ))
                            else: 
                                cands_to_use = []
                                
                            for k in range(min(req_hl - current_hl, len(cands_to_use))):
                                sid = cands_to_use[k]
                                schedule.at[sid, d_s] = '半/遅'
                                work_counts[sid] += 1
                                half_late_counts[sid] += 1
                                if is_sun_or_hol and i + 1 < len(dates):
                                    nd_s = dates[i+1].strftime('%Y-%m-%d')
                                    if pd.isna(schedule.at[sid, nd_s]): schedule.at[sid, nd_s] = 'OFF'

                    tgt_ids_weekday = [k for k,v in staff_info.items() if v['所属'] == '外来' and v['役職'] in ['副師長','看護師']]
                    pm_shifts = ['日', '遅', '半/遅', '年/遅', '年/', '半/', '特/', '年/半', '健/', '夏/', '研/', 'P11', 'P13', 'P17']
                    support_shifts = ['日', '遅', '半/遅', '年/遅', 'P4', 'P5', 'P8', 'P11', 'P13', 'P17', 'P70']

                    for i, d in enumerate(dates):
                        d_s = d.strftime('%Y-%m-%d')
                        is_hol = is_holiday_jp(d)
                        
                        pool_for_count = tgt_ids_holiday if is_hol else all_ns_ids
                        cands_pool_day = tgt_ids_holiday if is_hol else tgt_ids_weekday
                        
                        cur_am, cur_pm = 0, 0
                        for sid in pool_for_count:
                            val = schedule.at[sid, d_s]
                            if pd.notna(val):
                                is_outpatient = staff_info[sid]['所属'] == '外来'
                                if val in am_shifts and (is_outpatient or val in support_shifts): cur_am += 1
                                if val in pm_shifts and (is_outpatient or val in support_shifts): cur_pm += 1
                        req_am = req_dict[d_s]["am"]
                        req_pm = req_dict[d_s]["pm"]
                        
                        shortage_am = max(0, req_am - cur_am)
                        shortage_pm = max(0, req_pm - cur_pm)
                        needed_day = max(shortage_am, shortage_pm) 
                        
                        if needed_day > 0:
                            cands_day = []
                            for s in cands_pool_day: 
                                # 🌟パートさん（基本勤務が日以外）は【平日のみ】日勤のランダム抽選から外す
                                b_s = staff_info[s].get('基本勤務', '日')
                                if pd.isna(b_s) or b_s == '': b_s = '日'
                                if not is_hol and b_s != '日': continue
                                
                                if pd.isna(schedule.at[s, d_s]):
                                    is_prev_b = False
                                    if i > 0:
                                        p_s = dates[i-1].strftime('%Y-%m-%d')
                                        if schedule.at[s, p_s] == 'B': is_prev_b = True
                                    if not is_prev_b: cands_day.append(s)
                                    
                            if is_hol: 
                                cands_limit_ok = [s for s in cands_day if can_work_holiday_day(s, holiday_day_counts[s], staff_info)]
                                
                                cands_to_use = cands_limit_ok.copy()
                                random.shuffle(cands_to_use)
                                cands_to_use.sort(key=lambda x: (
                                    1 if is_consecutive_holiday_work(schedule, x, i, dates) else 0,
                                    -get_total_fixed_offs(x), 
                                    work_counts[x]
                                ))
                            else: 
                                cands_to_use = cands_day
                                random.shuffle(cands_to_use)
                                cands_to_use.sort(key=lambda x: work_counts[x])
                                
                            assigned_count = 0
                            for sid in cands_to_use:
                                if assigned_count >= needed_day:
                                    break
                                    
                                if is_hol and staff_info[sid]['所属'] == 'オペ室':
                                    current_op_count = sum(1 for check_sid in tgt_ids_holiday if schedule.at[check_sid, d_s] in am_shifts and staff_info[check_sid]['所属'] == 'オペ室')
                                    if current_op_count >= 1:
                                        continue 
                                        
                                schedule.at[sid, d_s] = '日'
                                work_counts[sid] += 1
                                if is_hol: holiday_day_counts[sid] += 1
                                assigned_count += 1

                    for sid in sids:
                        offs_1 = (schedule.loc[sid] == 'OFF').sum()
                        offs_05 = schedule.loc[sid].isin(['/半', '半/', '半/遅', '年/半', '健/半', '/夏', '夏/']).sum()
                        current_offs = offs_1 + (offs_05 * 0.5)
                        
                        empty_dates = schedule.columns[schedule.loc[sid].isna()].tolist()
                        empty_hols = [d for d in empty_dates if is_holiday_jp(datetime.datetime.strptime(d, '%Y-%m-%d').date())]
                        empty_weekdays = [d for d in empty_dates if not is_holiday_jp(datetime.datetime.strptime(d, '%Y-%m-%d').date())]
                        random.shuffle(empty_hols)
                        random.shuffle(empty_weekdays)
                        
                        needed_offs = max(0, target_offs_dict[sid] - current_offs)

                        for d_s in empty_hols:
                            schedule.at[sid, d_s] = 'OFF'
                            current_offs += 1
                            
                        for d_s in empty_weekdays:
                            if needed_offs >= 1:
                                schedule.at[sid, d_s] = 'OFF'
                                needed_offs -= 1
                            elif needed_offs == 0.5:
                                schedule.at[sid, d_s] = '/半' 
                                needed_offs -= 0.5
                            else:
                                d_obj = datetime.datetime.strptime(d_s, '%Y-%m-%d').date()
                                b_s = staff_info[sid].get('基本勤務', '日')
                                if pd.isna(b_s) or b_s == '': b_s = '日'
                                
                                # 🌟パートさんの空き枠には専用の基本勤務を入れる
                                if b_s != '日':
                                    schedule.at[sid, d_s] = b_s
                                else:
                                    default_shift = '日'
                                    if staff_info[sid]['所属'] == 'オペ室': default_shift = 'OP日'
                                    elif staff_info[sid]['所属'] == '内視鏡': default_shift = '内日'
                                    schedule.at[sid, d_s] = default_shift
                    
                    e_counts = {sid: 0 for sid in sids}
                    for d_s in cols:
                        if meta_e_color.at['E', d_s] == 'green':
                            e_name = e_schedule.at['E', d_s]
                            if e_name and e_name != "未定":
                                e_sids = staff_df[staff_df['名前'] == e_name]['職員番号'].values
                                if len(e_sids) > 0: e_counts[e_sids[0]] += 1
                                
                    for i, d_s in enumerate(cols):
                        if meta_e_color.at['E', d_s] != 'green':
                            super_cands = []     # 【最優先】日勤系 ＆ 翌日A
                            primary_cands = []   # 【第1候補】日勤系
                            secondary_cands = [] # 【第2候補】半/遅・年/遅
                            last_cands = []      # 【最終候補】B
                            
                            for s in sids:
                                e_limit = staff_info[s].get('待機上限', 0)
                                if e_limit > 0 and e_counts[s] < e_limit:
                                    current_shift = schedule.at[s, d_s]
                                    
                                    # 前日が当直だったかどうかのチェック
                                    is_post_tochoku = False
                                    if i > 0:
                                        prev_d_s = cols[i-1]
                                        if schedule.at[s, prev_d_s] == '当直':
                                            is_post_tochoku = True
                                            
                                    # 🌟【ここを追加】翌日が当直（当直前日）かどうかのチェック
                                    is_pre_tochoku = False
                                    if i + 1 < len(cols):
                                        next_d_s = cols[i+1]
                                        if schedule.at[s, next_d_s] == '当直':
                                            is_pre_tochoku = True
                                            
                                    # 🌟【さらに追加】前日または翌日が待機（E）かどうかのチェック
                                    is_adjacent_e = False
                                    s_name = staff_info[s]['名前']
                                    if i > 0 and e_schedule.at['E', cols[i-1]] == s_name:
                                        is_adjacent_e = True
                                    if i + 1 < len(cols) and e_schedule.at['E', cols[i+1]] == s_name:
                                        is_adjacent_e = True

                                    # 🌟【さらに追加】金曜待機の場合、土曜が日勤系ならブロックする
                                    is_fri_to_sat_day = False
                                    d_obj_for_e = datetime.datetime.strptime(d_s, '%Y-%m-%d').date()
                                    if d_obj_for_e.weekday() == 4 and i + 1 < len(cols):  # weekday() == 4 は金曜日
                                        next_s = cols[i+1]
                                        if schedule.at[s, next_s] in ['日', 'OP日', '内日']:
                                            is_fri_to_sat_day = True
                                            
                                    # 🌟【追加】翌日が「A（夜勤入り）」かどうかのチェック
                                    is_next_a = False
                                    if i + 1 < len(cols):
                                        next_s = cols[i+1]
                                        if schedule.at[s, next_s] == 'A':
                                            is_next_a = True

                                    # 🌟4段階の優先順位で振り分け（安全装置を全てクリアした人のみ）
                                    if not is_post_tochoku and not is_pre_tochoku and not is_adjacent_e and not is_fri_to_sat_day:
                                        if current_shift in ['日', 'OP日', '内日']:
                                            if is_next_a:
                                                super_cands.append(s)   # 【最優先】翌日A
                                            else:
                                                primary_cands.append(s) # 【第1候補】通常の日勤
                                        elif current_shift in ['半/遅', '年/遅']:
                                            secondary_cands.append(s)   # 【第2候補】
                                        elif current_shift == 'B':
                                            last_cands.append(s)        # 【最終候補】奥の手
                            
                            # 🌟優先順位の高いリストから採用する
                            if super_cands:
                                cands_to_use = super_cands
                            elif primary_cands:
                                cands_to_use = primary_cands
                            elif secondary_cands:
                                cands_to_use = secondary_cands
                            else:
                                cands_to_use = last_cands
                            
                            if cands_to_use:
                                random.shuffle(cands_to_use)
                                cands_to_use.sort(key=lambda x: e_counts[x])
                                chosen_e = cands_to_use[0]
                                e_schedule.at['E', d_s] = staff_info[chosen_e]['名前']
                                e_counts[chosen_e] += 1
                            else:
                                e_schedule.at['E', d_s] = "未定"
                                
                    st.session_state['generated_df'] = schedule
                    st.session_state['meta_color_df'] = meta_color
                    st.session_state['e_schedule_df'] = e_schedule
                    st.session_state['meta_e_color_df'] = meta_e_color
                    persist_current_schedule(selected_year, selected_month)
                    
                st.success("✅ 生成完了")
                st.rerun()

        comm_cells = set()
        for c in current_committees:
            if c['year'] == selected_year and c['month'] == selected_month:
                d_str = c['date']
                for sid in c['members']:
                    comm_cells.add((sid, d_str))

        if st.session_state['generated_df'] is not None:
            st.write("")
            
            roles_f = st.multiselect("役職フィルタ", ROLES, default=ROLES)
            groups_f = st.multiselect("所属フィルタ", GROUPS, default=GROUPS)
            
            filtered_staff = staff_df[(staff_df['役職'].isin(roles_f)) & (staff_df['所属'].isin(groups_f))].copy()
            group_order_map = {g: i for i, g in enumerate(GROUPS)}
            filtered_staff['所属順'] = filtered_staff['所属'].map(group_order_map)
            f_ids = filtered_staff.sort_values(['所属順', '表示順'])['職員番号'].values

            if not is_locked:
                # ★ ✨ 連休・一括入力パネル 
                st.write("### ✨ 連休・一括入力パネル")
                with st.container(border=True):
                    c_bulk1, c_bulk2, c_bulk3, c_bulk4, c_bulk5, c_bulk6 = st.columns([2, 2, 0.5, 2, 2, 2.5])
                    with c_bulk1:
                        bulk_staff_id = st.selectbox("👤 スタッフ", options=f_ids, format_func=lambda x: staff_df[staff_df['職員番号']==x]['名前'].iloc[0])
                    with c_bulk2:
                        bulk_start = st.selectbox("📅 開始日", options=date_strs, format_func=lambda x: col_map[x])
                    with c_bulk3:
                        st.markdown("<div style='text-align:center; padding-top:35px; font-weight:bold;'>〜</div>", unsafe_allow_html=True)
                    with c_bulk4:
                        bulk_end = st.selectbox("📅 終了日", options=date_strs, format_func=lambda x: col_map[x], index=len(date_strs)-1)
                    with c_bulk5:
                        bulk_shift = st.selectbox("📝 入力シフト", options=SHIFT_TYPES)
                    with c_bulk6:
                        exclude_holidays = st.checkbox("🌴 土日祝を除外", value=False)
                        if st.button("🚀 一括入力する", type="primary", use_container_width=True):
                            s_idx = date_strs.index(bulk_start)
                            e_idx = date_strs.index(bulk_end)
                            if s_idx > e_idx:
                                st.error("エラー：開始日は終了日より前に設定してください。")
                            else:
                                hist = st.session_state.get('schedule_history', [])
                                hist.append({
                                    'gen': st.session_state['generated_df'].copy(deep=True),
                                    'col': st.session_state['meta_color_df'].copy(deep=True),
                                    'e_gen': st.session_state['e_schedule_df'].copy(deep=True),
                                    'e_col': st.session_state['meta_e_color_df'].copy(deep=True)
                                })
                                if len(hist) > 4: hist.pop(0)
                                st.session_state['schedule_history'] = hist
                                
                                target_dates = date_strs[s_idx:e_idx+1]
                                for d in target_dates:
                                    if exclude_holidays and is_holiday_jp(datetime.datetime.strptime(d, '%Y-%m-%d').date()):
                                        continue
                                    st.session_state['generated_df'].at[bulk_staff_id, d] = bulk_shift
                                    st.session_state['meta_color_df'].at[bulk_staff_id, d] = 'green'
                                
                                persist_current_schedule(selected_year, selected_month)
                                st.success("一括入力が完了しました！")
                                st.session_state['editor_reset_key'] += 1
                                st.rerun()

            show_agg = st.checkbox("📊 右端に集計列（夜勤回数、休みの数など）を表示する", value=False)

            full_sch = st.session_state['generated_df']
            e_sch = st.session_state['e_schedule_df']
            color_ref = st.session_state['meta_color_df']
            e_color_ref = st.session_state['meta_e_color_df']
            
            f_ids = [sid for sid in f_ids if sid in full_sch.index]
            
            # (1) 右端集計データの計算
            agg_data = []
            for sid in f_ids:
                row = full_sch.loc[sid]
                night_cnt = int(row.isin(['A', '当直']).sum())
                late_cnt = int((row == '遅').sum())
                hl_cnt = int(row.isin(['半/遅', '年/遅']).sum())
                # ★ 代休はOFFに含めない
                off_cnt = float((row == 'OFF').sum() + row.isin(['/半', '半/', '半/遅', '年/半', '健/半', '/夏', '夏/']).sum() * 0.5)
                # ★ 年/遅を年休0.5として追加。代休はここにも含めない。
                paid_cnt = float((row == '年休').sum() + row.isin(['/年', '年/', '健/年', '年/半', '年/遅']).sum() * 0.5)
                e_cnt = 0
                if e_sch is not None:
                    s_name = staff_df[staff_df['職員番号']==sid]['名前'].iloc[0]
                    e_cnt = int((e_sch.loc['E'] == s_name).sum())
                off_str = str(int(off_cnt)) if off_cnt.is_integer() else f"{off_cnt:.1f}"
                paid_str = str(int(paid_cnt)) if paid_cnt.is_integer() else f"{paid_cnt:.1f}"
                agg_data.append({'職員番号': sid, '夜勤': str(night_cnt), '待機': str(e_cnt), '遅出': str(late_cnt), '半/遅': str(hl_cnt), 'OFF': off_str, '年休': paid_str})
            agg_df = pd.DataFrame(agg_data).set_index('職員番号')

            # (2) サマリー行の作成
            count_ns_ids = [k for k,v in staff_df.set_index("職員番号").to_dict('index').items() if v['役職'] in ['副師長','看護師']]
            count_all_ids = [k for k,v in staff_df.set_index("職員番号").to_dict('index').items() if v['役職'] in ['師長', '副師長','看護師']]
            am_shifts = ['日', '/年', '/半', '/特', '/夏', '/研', 'P4', 'P5', 'P8', 'P11', 'P13', 'P17', 'P70']
            pm_shifts = ['日', '遅', '半/遅', '年/遅', '年/', '半/', '特/', '年/半', '健/', '夏/', '研/', 'P11', 'P13', 'P17']
            support_shifts = ['日', '遅', '半/遅', '年/遅', 'P4', 'P5', 'P8', 'P11', 'P13', 'P17', 'P70']
            staff_group_dict = staff_df.set_index("職員番号")["所属"].to_dict()

            rows = []
            for n in ["AM", "PM", "待機"]:
                r = {"名前": n, "所属": "集計", "役職": ""}
                for ds in date_strs:
                    if n == "待機":
                        val = e_sch.at['E', ds]
                        c = e_color_ref.at['E', ds] if ds in e_color_ref.columns else 'black'
                        val_str = str(val) if pd.notna(val) else ""
                        r[col_map[ds]] = f"{val_str} 🟢" if c == 'green' and val_str and val_str != "未定" else val_str
                    else:
                        d_obj = datetime.datetime.strptime(ds, '%Y-%m-%d').date()
                        pool = count_all_ids if is_holiday_jp(d_obj) else count_ns_ids
                        target_pool = [s for s in pool if s in full_sch.index]
                        
                        count = 0
                        s_list = am_shifts if n == "AM" else pm_shifts
                        for sid in target_pool:
                            val = full_sch.at[sid, ds]
                            if pd.notna(val) and val in s_list:
                                if staff_group_dict.get(sid, '外来') == '外来' or val in support_shifts:
                                    count += 1
                        
                        r[col_map[ds]] = str(count)
                rows.append(r)

            # 🌟【削除】ここにあった「rows.append({"名前": "━━━━━━━━━━━━━━━━" ...」の行を丸ごと消します！
            
            # (3) スタッフ行の作成（グループごと）
            name_counts = staff_df['名前'].value_counts()
            for grp in GROUPS:
                if grp in groups_f:
                    # 🌟【修正】線を短くスッキリさせます！
                    rows.append({"名前": f"━ {grp} ━", "所属": "区切り", "役職": ""})
                    grp_staff = filtered_staff[filtered_staff['所属'] == grp].sort_values('表示順')
                    for sid in grp_staff['職員番号']:
                        if sid not in full_sch.index: continue
                        raw_name = grp_staff.set_index("職員番号").at[sid, "名前"]
                        disp_name = f"{raw_name} ({sid})" if name_counts.get(raw_name, 0) > 1 else raw_name
                        s_row = {"名前": disp_name, "所属": grp, "役職": grp_staff.set_index("職員番号").at[sid, "役職"]}
                        
                        for ds in date_strs:
                            val = full_sch.at[sid, ds]
                            c = color_ref.at[sid, ds] if sid in color_ref.index and ds in color_ref.columns else 'black'
                            is_comm = (sid, ds) in comm_cells
                            v_str = str(val) if pd.notna(val) else ""
                            if c == 'red': v_str += " 🔴"
                            elif c == 'blue': v_str += " 🔵"
                            elif c == 'green': v_str += " 🟢"
                            elif is_comm: v_str += " 🟡"
                            s_row[col_map[ds]] = v_str
                            
                        if show_agg:
                            for k in ['夜勤', '待機', '遅出', '半/遅', 'OFF', '年休']:
                                s_row[k] = agg_df.at[sid, k]
                                
                        rows.append(s_row)

            # (4) 統合データフレーム化とカラム設定
            all_view = pd.DataFrame(rows).set_index("名前")
            if show_agg:
                for k in ['夜勤', '待機', '遅出', '半/遅', 'OFF', '年休']:
                    if k not in all_view.columns: all_view[k] = "" 

            col_cfg = {
                "所属": None,
                "役職": None
            }
            # 🌟【修正】待機上限が1以上、または役職が「師長」の人の名前をリストに入れる
            e_staff_names = ["", "未定"] + staff_df[(staff_df['待機上限'] > 0) | (staff_df['役職'] == '師長')]['名前'].tolist()
            special_leaves = ['介護休暇', '病欠', '産休', '育休']
            main_options = [""] + [st_type for st_type in SHIFT_TYPES if st_type not in special_leaves]
            
            for s in short_dates:
                current_vals = all_view[s].dropna().astype(str).unique().tolist()
                comb_options = list(dict.fromkeys(main_options + special_leaves + e_staff_names + current_vals))
                col_cfg[s] = st.column_config.SelectboxColumn(width="small", options=comb_options)
            
            if show_agg:
                for k in ['夜勤', '待機', '遅出', '半/遅', 'OFF', '年休']:
                    col_cfg[k] = st.column_config.TextColumn(disabled=True, width="small")

            # (5) 色付けスタイル
            def styled_all_view(df):
                styles = pd.DataFrame('', index=df.index, columns=df.columns)
                for col in df.columns:
                    is_holiday_col = False
                    is_sat = False
                    if col in rev_col_map:
                        d_obj = datetime.datetime.strptime(rev_col_map[col], '%Y-%m-%d').date()
                        is_holiday_col = (d_obj.weekday() == 6 or jpholiday.is_holiday(d_obj))
                        is_sat = (d_obj.weekday() == 5 and not is_holiday_col)
                        
                    for idx in df.index:
                        bg = ""
                        # 🌟【修正】消した行の条件を外し、短い線「━」に条件を合わせます！
                        if idx in ["AM", "PM"] or "━" in str(idx):
                            bg = "background-color: #f0f2f6; font-weight: bold;"
                        elif idx == "待機":
                            bg = "background-color: #fff2cc; font-weight: bold;"
                        else:
                            if col in rev_col_map:
                                if is_holiday_col: bg = "background-color: #ffe6e6; font-weight: bold;"
                                elif is_sat: bg = "background-color: #e6f2ff; font-weight: bold;"
                                else: bg = "font-weight: bold;"
                            else:
                                bg = "background-color: #f8f9fa;" # 集計列
                        styles.at[idx, col] = bg
                return styles

            # (6) 画面描画
            st.markdown("<span style='font-weight:bold;'>【凡例】</span> <span style='color:#d32f2f; font-weight:bold;'>🔴 赤希望</span>　<span style='color:#1976d2; font-weight:bold;'>🔵 青希望</span>　<span style='color:#388e3c; font-weight:bold;'>🟢 師長確定</span>　<span style='color:#fbc02d; font-weight:bold;'>🟡 委員会</span>", unsafe_allow_html=True)
            st.info("💡 下の表は【サマリー・待機・シフト】が全て合体しています。編集後は必ず一番下の「一括保存」ボタンを押してください。")
            
            edited_all = st.data_editor(
                all_view.style.apply(styled_all_view, axis=None), 
                column_config=col_cfg, 
                use_container_width=True, 
                height=min((len(all_view)+1)*36, 1200),
                disabled=is_locked, # ★ ここで閲覧専用モードを適用
                key=f"unified_editor_{st.session_state['editor_reset_key']}"
            )

            # --- 7. 一括保存ロジック ---
            if not is_locked:
                st.write("---")
                st.subheader("💾 シフトの確定保存")
                
                # 保存ボタンと穴埋めボタンを横並びに配置
                c_save1, c_save2 = st.columns(2)
                with c_save1:
                    if st.button("🔄 編集内容を確定して一括保存する", type="primary", use_container_width=True):
                        # 判定用の関数（空欄やNoneを統一して扱う）
                        def clean_val(v):
                            if pd.isna(v) or v is None or str(v).lower() in ['nan', 'none', '']:
                                return ""
                            return str(v).replace(" 🟢", "").replace(" 🔴", "").replace(" 🔵", "").replace(" 🟡", "").strip()

                        # 作業用のコピーを作成
                        temp_e_gen = st.session_state['e_schedule_df'].copy()
                        temp_e_col = st.session_state['meta_e_color_df'].copy()
                        temp_gen = st.session_state['generated_df'].copy()
                        temp_col = st.session_state['meta_color_df'].copy()
                        
                        e_changed = False
                        shifts_changed = False
                        sensitive_changes = []
                        
                        # 待機行の判定
                        e_row_new = edited_all.loc["待機"]
                        for sd, ds in rev_col_map.items():
                            new_val = clean_val(e_row_new[sd])
                            old_val = clean_val(temp_e_gen.at['E', ds])
                            
                            if new_val != old_val:
                                e_changed = True
                                temp_e_gen.at['E', ds] = new_val if new_val != "" else np.nan
                                temp_e_col.at['E', ds] = 'green'

                        # スタッフ行の判定
                        for name_idx, row in edited_all.iterrows():
                            # 🌟【修正】線が1本（━）でも区切り線としてちゃんと無視するようにする
                            if "━" in str(name_idx) or name_idx in ["AM", "PM", "待機"]: continue
                            sid = int(name_idx.split('(')[-1].split(')')[0]) if '(' in name_idx else staff_df[staff_df['名前']==name_idx]['職員番号'].iloc[0]
                            if sid not in temp_gen.index: continue
                            
                            for sd, ds in rev_col_map.items():
                                new_val = clean_val(row[sd])
                                old_val = clean_val(temp_gen.at[sid, ds])
                                
                                if new_val != old_val:
                                    # ★ 赤・青希望の書き換えをキャッチ！
                                    original_color = st.session_state['meta_color_df'].at[sid, ds]
                                    if original_color in ['red', 'blue'] and new_val != old_val:
                                        raw_name = staff_df[staff_df['職員番号']==sid]['名前'].iloc[0]
                                        sensitive_changes.append({
                                            'idx': sid, 'col': ds, 'name': raw_name,
                                            'date_label': col_map[ds], 'old': old_val, 'new': new_val, 'color': original_color
                                        })
                                        
                                    shifts_changed = True
                                    temp_gen.at[sid, ds] = new_val if new_val != "" else np.nan
                                    temp_col.at[sid, ds] = 'green'
                                    
                                    # 🌟 A(入り) -> B(明け) -> OFF(休み) の自動連動ロジック 🌟
                                    if new_val == "A":
                                        idx = date_strs.index(ds)
                                        if idx + 1 < len(date_strs):
                                            next_ds1 = date_strs[idx + 1]
                                            temp_gen.at[sid, next_ds1] = "B"
                                            temp_col.at[sid, next_ds1] = 'green'
                                        if idx + 2 < len(date_strs):
                                            next_ds2 = date_strs[idx + 2]
                                            temp_gen.at[sid, next_ds2] = "OFF"
                                            temp_col.at[sid, next_ds2] = 'green'

                                    # 🌟 日曜日の「半/遅・年/遅」 -> 翌日OFF の自動連動ロジック 🌟
                                    if new_val in ["半/遅", "年/遅"]:
                                        d_obj = datetime.datetime.strptime(ds, '%Y-%m-%d').date()
                                        if d_obj.weekday() == 6:
                                            idx = date_strs.index(ds)
                                            if idx + 1 < len(date_strs):
                                                next_ds1 = date_strs[idx + 1]
                                                temp_gen.at[sid, next_ds1] = "OFF"
                                                temp_col.at[sid, next_ds1] = 'green'

                        if e_changed or shifts_changed:
                            if sensitive_changes:
                                # 希望を上書きしている場合はアラート画面を表示（保存は保留）
                                confirm_sensitive_changes(sensitive_changes, temp_gen, temp_col, e_changed, temp_e_gen, temp_e_col, selected_year, selected_month)
                            else:
                                # 問題なければそのまま保存
                                hist = st.session_state.get('schedule_history', [])
                                hist.append({
                                    'gen': st.session_state['generated_df'].copy(deep=True),
                                    'col': st.session_state['meta_color_df'].copy(deep=True),
                                    'e_gen': st.session_state['e_schedule_df'].copy(deep=True),
                                    'e_col': st.session_state['meta_e_color_df'].copy(deep=True)
                                })
                                if len(hist) > 4: hist.pop(0)
                                st.session_state['schedule_history'] = hist
                                
                                st.session_state['generated_df'] = temp_gen
                                st.session_state['meta_color_df'] = temp_col
                                st.session_state['e_schedule_df'] = temp_e_gen
                                st.session_state['meta_e_color_df'] = temp_e_col
                                
                                persist_current_schedule(selected_year, selected_month)
                                st.success("✅ シフトを確定保存しました！")
                                st.rerun()

                with c_save2:
                    if st.button("🎯 ピンポイント穴埋めアシスト", use_container_width=True):
                        pinpoint_fill_dialog(date_strs, col_map, staff_df, current_wishes, selected_year, selected_month)

                st.write("---")
                st.write("### 🔓 手動確定(🟢)の解除")
                st.caption("手動変更でロックされたコマを選択して解除できます。解除すると、次回の自動生成で再計算の対象になります。")
                
                green_cells = [{"type": "E", "date_col": col_d, "label": f"待機 - {col_map[col_d]} (現在: {st.session_state['e_schedule_df'].at['E', col_d]})"} for col_d in date_strs if col_d in st.session_state['meta_e_color_df'].columns and st.session_state['meta_e_color_df'].at['E', col_d] == 'green']
                green_cells += [{"type": "shift", "sid": sid, "date_col": col_d, "label": f"シフト: {staff_df[staff_df['職員番号']==sid]['名前'].iloc[0]} - {col_map[col_d]} (現在: {st.session_state['generated_df'].at[sid, col_d]})"} for sid in f_ids if sid in st.session_state['meta_color_df'].index for col_d in date_strs if col_d in st.session_state['meta_color_df'].columns and st.session_state['meta_color_df'].at[sid, col_d] == 'green']
                
                if green_cells:
                    unlock_targets = st.multiselect("解除するコマを選択してください", options=[g["label"] for g in green_cells])
                    if st.button("選択したコマのロックを解除"):
                        hist = st.session_state.get('schedule_history', [])
                        hist.append({
                            'gen': st.session_state['generated_df'].copy(deep=True),
                            'col': st.session_state['meta_color_df'].copy(deep=True),
                            'e_gen': st.session_state['e_schedule_df'].copy(deep=True),
                            'e_col': st.session_state['meta_e_color_df'].copy(deep=True)
                        })
                        if len(hist) > 4: hist.pop(0)
                        st.session_state['schedule_history'] = hist
                        
                        for label in unlock_targets:
                            target = next(g for g in green_cells if g["label"] == label)
                            if target["type"] == "E": st.session_state['meta_e_color_df'].at['E', target["date_col"]] = 'black'
                            else:
                                original_color, original_shift = next(((w['type'], w['shift']) for w in current_wishes if w['staff_id'] == target["sid"] and w['date'] == target["date_col"]), ('black', None))
                                st.session_state['meta_color_df'].at[target["sid"], target["date_col"]] = original_color
                                if original_shift is not None: st.session_state['generated_df'].at[target["sid"], target["date_col"]] = original_shift
                        persist_current_schedule(selected_year, selected_month)
                        st.session_state['editor_reset_key'] += 1
                        st.rerun()
                else:
                    st.info("現在、ロック（確定）されているコマはありません。")

            # ==========================================
            # ★ リアルタイム ルール整合性チェック
            # ==========================================
            st.markdown("---")
            st.subheader("🚨 リアルタイム ルール整合性チェック")
            
            fin_sch = st.session_state['generated_df'].copy()
            fin_e = st.session_state['e_schedule_df'].copy()
            errs = []
            
            for d_s in date_strs:
                for sid in f_ids:
                    if sid in fin_sch.index:
                        val = fin_sch.at[sid, d_s]
                        if pd.isna(val) or str(val).strip() == "" or str(val) == "---":
                            staff_name = staff_df[staff_df['職員番号']==sid]['名前'].iloc[0]
                            errs.append(f"❌ {col_map[d_s]}: {staff_name} さんのシフトが空欄（未入力）です！")
            
            libre_cws = staff_df[staff_df['リブレ'] == True]['職員番号'].tolist()
            dm_ns = staff_df[staff_df['DM'] == True]['職員番号'].tolist()

            for d_s in date_strs:
                d_obj = datetime.datetime.strptime(d_s, '%Y-%m-%d').date()
                wd = d_obj.weekday()
                is_hol = is_holiday_jp(d_obj)
                
                closed_types_today = [c['clinic_type'] for c in current_closed_days if c['date'] == d_s]
                
                if not is_hol:
                    if wd in [1, 3]: 
                        if 'リブレ' not in closed_types_today:
                            if libre_cws and not any(fin_sch.at[cw, d_s] in am_shifts for cw in libre_cws if cw in fin_sch.index):
                                errs.append(f"❌ {col_map[d_s]}: リブレCWが日勤帯にいません！")
                        if 'DM' not in closed_types_today:
                            if dm_ns and not any(fin_sch.at[ns, d_s] in am_shifts for ns in dm_ns if ns in fin_sch.index):
                                errs.append(f"💡 {col_map[d_s]}: DM看護師が日勤帯にいません（他部署からの応援が必要です）")

            e_counts_check = fin_e.loc['E'].value_counts()
            for ename, count in e_counts_check.items():
                if pd.notna(ename) and ename != '未定' and ename != "":
                    e_sids = staff_df[staff_df['名前'] == ename]
                    if not e_sids.empty:
                        limit = e_sids['待機上限'].values[0]
                        if count > limit: errs.append(f"⚠️ {ename}さんの待機回数({count}回)が上限({limit}回)を超えています")
            
            for c in current_committees:
                if c['year'] == selected_year and c['month'] == selected_month:
                    d_s = c['date']
                    if d_s in date_strs:
                        for sid in c['members']:
                            if sid in fin_sch.index:
                                actual_shift = fin_sch.at[sid, d_s]
                                staff_name = staff_df[staff_df['職員番号']==sid]['名前'].iloc[0]
                                wished_off = any(w['staff_id'] == sid and w['date'] == d_s and w['shift'] in ['OFF', '年休', '特休', '夏休', '/年', '年/', '/半', '半/', '年/半', '/特', '特/', '/夏', '夏/', '介護休暇', '病欠', '産休', '育休', '代休', '/代休'] for w in current_wishes)
                                
                                if wished_off: errs.append(f"⚠️ {col_map[d_s]}: {staff_name}さんは休み希望ですが「{c['committee']}」が入っています")
                                if actual_shift in ['OFF', '年休', '特休', '夏休', '/年', '年/', '/半', '半/', '年/半', '/特', '特/', '/夏', '夏/', '健/半', '健/年', '健/', '介護休暇', '病欠', '産休', '育休', '代休', '/代休'] and not wished_off:
                                    errs.append(f"⚠️ {col_map[d_s]}: {staff_name}さんは「{c['committee']}」の予定ですが、お休み系または健診シフト({actual_shift})になっています")
                                elif actual_shift in ['遅', '半/遅', '年/遅', 'A', 'B', '当直']:
                                    errs.append(f"💡 {col_map[d_s]}: {staff_name}さんは「{c['committee']}」の予定ですが、シフトが「{actual_shift}」になっています。（参加できるか確認してください）")

            # ★ 学生判定対応の公休数チェック
            num_holidays = sum(1 for d in dates if is_holiday_jp(d))
            sun_hol_count = sum(1 for d in dates if d.weekday() == 6 or jpholiday.is_holiday(d))
            
            for sid in f_ids:
                if sid in fin_sch.index:
                    offs_1 = (fin_sch.loc[sid] == 'OFF').sum()
                    offs_05 = fin_sch.loc[sid].isin(['/半', '半/', '半/遅', '年/半', '健/半', '/夏', '夏/']).sum()
                    total_offs = offs_1 + (offs_05 * 0.5)
                    
                    is_student = staff_df[staff_df['職員番号']==sid]['学生'].iloc[0] if '学生' in staff_df.columns else False
                    target_offs = sun_hol_count if is_student else num_holidays
                    
                    if total_offs != target_offs:
                        staff_name = staff_df[staff_df['職員番号']==sid]['名前'].iloc[0]
                        errs.append(f"❌ 【公休数】 {staff_name}さんの公休(OFF等)が {total_offs}日 になっています（今月の規定: {target_offs}日）")

            for d_s in date_strs:
                d_obj = datetime.datetime.strptime(d_s, '%Y-%m-%d').date()
                is_sun_or_hol = d_obj.weekday() == 6 or is_holiday_jp(d_obj)
                is_sat_only = d_obj.weekday() == 5 and not is_holiday_jp(d_obj)
                
                # ★修正：マスタにはいるが現在の表にはいないスタッフを除外する
                if is_holiday_jp(d_obj):
                    valid_ids = [sid for sid in count_all_ids if sid in fin_sch.index]
                else:
                    valid_ids = [sid for sid in count_ns_ids if sid in fin_sch.index]
                
                if is_holiday_jp(d_obj):
                    for s in fin_sch.index:
                        val = fin_sch.at[s, d_s]
                        if val in ['OP日', '内日', '内P7', '内P9', '内P16']:
                            staff_name = staff_df[staff_df['職員番号']==s]['名前'].iloc[0]
                            errs.append(f"❌ {col_map[d_s]}: 土日祝ですが、{staff_name}さんに「{val}」が入力されています（休日は「日」に変更してください）")

                support_shifts = ['日', '遅', '半/遅', '年/遅', 'P4', 'P5', 'P8', 'P11', 'P13', 'P17', 'P70']
                real_am, real_pm, real_late, real_hl = 0, 0, 0, 0
                for sid in valid_ids:
                    val = fin_sch.at[sid, d_s]
                    if pd.notna(val):
                        is_out = staff_df[staff_df['職員番号']==sid]['所属'].iloc[0] == '外来'
                        if val in am_shifts and (is_out or val in support_shifts): real_am += 1
                        if val in pm_shifts and (is_out or val in support_shifts): real_pm += 1
                        if val == '遅' and (is_out or val in support_shifts): real_late += 1
                        if val in ['半/遅', '年/遅'] and (is_out or val in support_shifts): real_hl += 1
                req_am = req_dict[d_s]["am"]
                req_pm = req_dict[d_s]["pm"]
                req_late = req_dict[d_s]["late"]
                req_hl = req_dict[d_s]["half_late"]
                
                if is_sun_or_hol:
                    if real_am != req_am: errs.append(f"❌ {col_map[d_s]}: 日曜・祝日の日勤人数が絶対条件({req_am}人)と異なります（現在 {real_am}人）")
                    if real_hl != req_hl: errs.append(f"❌ {col_map[d_s]}: 日曜・祝日の半/遅・年/遅人数が絶対条件({req_hl}人)と異なります（現在 {real_hl}人）")
                    if real_pm < req_pm: errs.append(f"💡 {col_map[d_s]}: PM不足 ({real_pm}/{req_pm})")
                elif is_sat_only:
                    if real_am != req_am: errs.append(f"❌ {col_map[d_s]}: 土曜日の日勤人数が絶対条件({req_am}人)と異なります（現在 {real_am}人）")
                    if real_late != req_late: errs.append(f"❌ {col_map[d_s]}: 土曜日の遅出人数が絶対条件({req_late}人)と異なります（現在 {real_late}人）")
                    if real_pm < req_pm: errs.append(f"💡 {col_map[d_s]}: PM不足 ({real_pm}/{req_pm})")
                else:
                    if real_am < req_am: errs.append(f"💡 {col_map[d_s]}: AM不足 ({real_am}/{req_am})")
                    if real_pm < req_pm: errs.append(f"💡 {col_map[d_s]}: PM不足 ({real_pm}/{req_pm})")
                    if real_late < req_late: errs.append(f"💡 {col_map[d_s]}: 遅出不足 ({real_late}/{req_late})")
                    if real_hl < req_hl: errs.append(f"💡 {col_map[d_s]}: 半/遅・年/遅不足 ({real_hl}/{req_hl})")

                # ★ 休日の日勤におけるオペ室スタッフの人数をチェック
                if is_holiday_jp(d_obj):
                    op_day_count = sum(1 for s in valid_ids if fin_sch.at[s, d_s] in am_shifts and staff_df[staff_df['職員番号']==s]['所属'].iloc[0] == 'オペ室')
                    if op_day_count >= 2:
                        errs.append(f"⚠️ {col_map[d_s]}: 休日の日勤にオペ室所属のスタッフが2人以上割り当てられています（現在 {op_day_count}人）")
                    
                if (fin_sch[d_s] == 'A').sum() != 1: errs.append(f"❌ {col_map[d_s]}: 夜勤(A)人数エラー")
                    
                if is_sat_only:
                    if d_s in date_strs[:-1]: 
                        next_d_s = date_strs[date_strs.index(d_s) + 1]
                        for s in fin_sch.index: 
                            if fin_sch.at[s, d_s] == '遅':
                                n_val = fin_sch.at[s, next_d_s]
                                if pd.notna(n_val) and n_val not in REST_AND_A_SHIFTS and n_val != "" and n_val != "---":
                                    staff_name = staff_df[staff_df['職員番号']==s]['名前'].iloc[0]
                                    errs.append(f"⚠️ {col_map[d_s]}: {staff_name}さんの遅出の翌日が「{n_val}」になっています (お休み系かAであるべきです)")
                                    
                if is_sun_or_hol:
                    if d_s in date_strs[:-1]: 
                        next_d_s = date_strs[date_strs.index(d_s) + 1]
                        for s in fin_sch.index: 
                            if fin_sch.at[s, d_s] in ['半/遅', '年/遅']:
                                n_val = fin_sch.at[s, next_d_s]
                                if pd.notna(n_val) and n_val not in REST_AND_A_SHIFTS and n_val != "" and n_val != "---":
                                    staff_name = staff_df[staff_df['職員番号']==s]['名前'].iloc[0]
                                    errs.append(f"⚠️ {col_map[d_s]}: {staff_name}さんの半/遅・年/遅の翌日が「{n_val}」になっています (お休み系かAであるべきです)")
                                    
                e_name = fin_e.at['E', d_s]
                if pd.isna(e_name) or str(e_name).strip() == "" or e_name == "未定":
                    errs.append(f"❌ {col_map[d_s]}: 待機の担当者が空欄（未定）です！")
                else:
                    e_sids = staff_df[staff_df['名前'] == e_name]['職員番号'].values
                    if len(e_sids) > 0:
                        sid = e_sids[0]
                        shift = fin_sch.at[sid, d_s]
                        if shift not in E_CAPABLE_SHIFTS: errs.append(f"❌ {col_map[d_s]}: 待機担当の {e_name} さんが対応できないシフト「{shift}」になっています")
                        # 🌟【修正】師長なら待機上限0でもエラーを出さない（特例扱い）
                        staff_info_check = staff_df[staff_df['職員番号'] == sid].iloc[0]
                        if staff_info_check['待機上限'] == 0 and staff_info_check['役職'] != '師長':
                            errs.append(f"❌ {col_map[d_s]}: 待機上限0の {e_name} さんが待機に割り当てられています")

            # 🌟【ここから追加】6連勤以上（日勤系）のチェック🌟
            # 夜勤（A, B, 当直）や 休み（OFFなど）以外の実働シフトを「日勤系」とみなす
            non_day_shifts = ['A', 'B', '当直', 'OFF', '年休', '特休', '夏休', '/年', '年/', '/半', '半/', '年/半', '/特', '特/', '/夏', '夏/', '代休', '/代休', '健/年', '健/半', '健/', '---', '介護休暇', '病欠', '産休', '育休']
            
            for sid in f_ids:
                if sid in fin_sch.index:
                    staff_name = staff_df[staff_df['職員番号']==sid]['名前'].iloc[0]
                    consecutive_days = 0
                    start_d_s = ""
                    
                    for d_s in date_strs:
                        val = fin_sch.at[sid, d_s]
                        if pd.notna(val) and str(val).strip() != "":
                            if val not in non_day_shifts:
                                if consecutive_days == 0:
                                    start_d_s = d_s
                                consecutive_days += 1
                            else:
                                if consecutive_days >= 6:
                                    start_label = col_map[start_d_s]
                                    errs.append(f"⚠️ {staff_name}さんが {start_label} から日勤系のシフトで {consecutive_days}連勤 になっています")
                                consecutive_days = 0
                        else:
                            if consecutive_days >= 6:
                                start_label = col_map[start_d_s]
                                errs.append(f"⚠️ {staff_name}さんが {start_label} から日勤系のシフトで {consecutive_days}連勤 になっています")
                            consecutive_days = 0
                            
                    if consecutive_days >= 6:
                        start_label = col_map[start_d_s]
                        errs.append(f"⚠️ {staff_name}さんが {start_label} から日勤系のシフトで {consecutive_days}連勤 になっています")
            # 🌟【ここまで追加】🌟

            def get_err_priority(msg):
                if msg.startswith("❌"):
                    if "【公休数】" in msg:
                        return 2  
                    return 1      
                if msg.startswith("⚠️"): return 3
                if msg.startswith("💡"): return 4
                return 5

            errs.sort(key=get_err_priority)

            if not errs:
                st.success("✅ 全てのルールをクリアしています！パーフェクトです！")
            else:
                err_x = [e for e in errs if e.startswith("❌")]
                err_w = [e for e in errs if e.startswith("⚠️")]
                err_i = [e for e in errs if e.startswith("💡")]
                
                st.write(f"**現在、{len(errs)}件のチェック結果があります**")
                
                if err_x:
                    with st.expander(f"❌ エラー（絶対に直すべき項目）: {len(err_x)}件", expanded=True):
                        for e in err_x:
                            st.error(e)
                
                if err_w:
                    with st.expander(f"⚠️ 警告（できれば直したい項目）: {len(err_w)}件", expanded=False):
                        for e in err_w:
                            st.warning(e)
                            
                if err_i:
                    with st.expander(f"💡 注意（確認だけでOKな項目・人数不足など）: {len(err_i)}件", expanded=False):
                        for e in err_i:
                            st.info(e)

            # ==========================================
            # ★ 8. シフト表のダウンロード（最終形態Excel / CSV）
            # ==========================================
            st.markdown("---")
            st.subheader("📥 シフト表のダウンロード")
            
            import io
            from openpyxl.styles import Font, Alignment
            
            c_dl1, c_dl2 = st.columns(2)
            
            # (1) 従来のCSVダウンロード
            out_csv = all_view.reset_index().drop(columns=['夜勤', '待機', '遅出', '半/遅', 'OFF', '年休', '所属', '役職'], errors='ignore')
            out_csv = out_csv[~out_csv['名前'].str.contains("━")]
            with c_dl1:
                st.download_button("📥 CSV形式でダウンロード", out_csv.to_csv(index=False).encode('utf-8_sig'), "shift_data.csv", use_container_width=True)
                
            # (2) 提出用Excelダウンロード
            with c_dl2:
                # 和暦への変換
                reiwa_year = selected_year - 2018
                era_str = f"令和{reiwa_year}年{selected_month}月"
                
                # 委員会と待機(E)のマッピング
                comm_map = {}
                for c in current_committees:
                    if c['year'] == selected_year and c['month'] == selected_month:
                        d_str = c['date']
                        for sid in c['members']:
                            comm_map[(sid, d_str)] = c['committee']
                
                # --- データの構築 ---
                ex_data = []
                
                # 行1: タイトル（後でマージして入力するため空行）
                ex_data.append([""] * (2 + len(dates) + 6))

                # 行2: 日付行 (1, 2, 3...)
                date_row = ["", "日付"] + [d.day for d in dates] + [""] * 6
                ex_data.append(date_row)

                # 行3: 曜日行 (月, 火, 水...)
                wd_names = ['月','火','水','木','金','土','日']
                weekday_row = ["所属", "氏名／曜日"] + [wd_names[d.weekday()] for d in dates] + ['夜勤', '待機', '遅出', '半/遅', 'OFF', '年休']
                ex_data.append(weekday_row)

                # 行4-6: 空白行（委員会・医師休みなど）
                ex_data.append(["", "委員会"] + [""] * (len(dates) + 6))
                ex_data.append(["", "医師休み"] + [""] * (len(dates) + 6))
                ex_data.append(["", "透防・ケモ・神"] + [""] * (len(dates) + 6))
                
                # 行7-8: 日勤数 (AM / PM)
                summary_rows = [("日勤数(AM)", am_shifts), ("日勤数(PM)", pm_shifts)]
                support_shifts = ['日', '遅', '半/遅', '年/遅', 'P4', 'P5', 'P8', 'P11', 'P13', 'P17', 'P70']
                staff_group_dict = staff_df.set_index("職員番号")["所属"].to_dict()
                
                for label, s_list in summary_rows:
                    row_data = ["", "", label, ""]
                    for ds in date_strs:
                        d_obj = datetime.datetime.strptime(ds, '%Y-%m-%d').date()
                        is_hol = is_holiday_jp(d_obj) or d_obj.weekday() == 6
                        pool = count_all_ids if is_hol else count_ns_ids
                        target_pool = [sid for sid in pool if sid in full_sch.index]
                        
                        count = 0
                        for sid in target_pool:
                            val = full_sch.at[sid, ds]
                            if pd.notna(val) and val in s_list:
                                if staff_group_dict.get(sid, '外来') == '外来' or val in support_shifts:
                                    count += 1
                                    
                        row_data.append(str(count))
                    row_data += [""] * 6
                    ex_data.append(row_data)

                # --- スタッフ行の作成 ---
                cell_styles = {} 
                staff_start_row_idx = len(ex_data)
                curr_row_idx = staff_start_row_idx
                
                for g_idx, grp in enumerate(GROUPS):
                    if grp not in groups_f: continue
                    if g_idx > 0:
                        ex_data.append([""] * len(weekday_row))
                        curr_row_idx += 1
                        
                    grp_staff = filtered_staff[filtered_staff['所属'] == grp].sort_values('表示順')
                    for sid in grp_staff['職員番号']:
                        if sid not in full_sch.index: continue
                        
                        raw_name = grp_staff.set_index("職員番号").at[sid, "名前"]
                        ex_row = [grp, raw_name] 
                        
                        for c_idx, ds in enumerate(date_strs):
                            val = full_sch.at[sid, ds]
                            val = str(val) if pd.notna(val) else ""
                            d_obj = datetime.datetime.strptime(ds, '%Y-%m-%d').date()
                            is_hol = is_holiday_jp(d_obj) or d_obj.weekday() == 6
                            
                            if val in ['日', 'OP日', '内日']:
                                val = "日" if is_hol else ""
                            if (sid, ds) in comm_map: val = comm_map[(sid, ds)]
                            if e_sch is not None:
                                if str(e_sch.at['E', ds]).strip() == str(raw_name).strip():
                                    val = (str(val) + "E").strip()
                            ex_row.append(val)
                            
                            c_ref = color_ref.at[sid, ds] if sid in color_ref.index and ds in color_ref.columns else 'black'
                            if c_ref == 'red': cell_styles[(curr_row_idx, c_idx + 2)] = "FF0000"
                            elif c_ref == 'blue': cell_styles[(curr_row_idx, c_idx + 2)] = "0000FF"
                                
                        for agg in ['夜勤', '待機', '遅出', '半/遅', 'OFF', '年休']:
                            ex_row.append(agg_df.at[sid, agg] if sid in agg_df.index else "")
                        
                        ex_data.append(ex_row)
                        curr_row_idx += 1

                # --- Excel保存と装飾 ---
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    pd.DataFrame(ex_data).to_excel(writer, index=False, header=False, sheet_name='シフト表')
                    ws = writer.sheets['シフト表']
                    
                    from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
                    thin_border = Border(left=Side(style='thin', color='000000'), 
                                         right=Side(style='thin', color='000000'), 
                                         top=Side(style='thin', color='000000'), 
                                         bottom=Side(style='thin', color='000000'))
                    
                    # 目に優しいパステル調の土日祝背景色
                    sat_fill = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid') 
                    hol_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid') 
                    
                    # 🌟【修正】誰のシフトにも釣られず、その月の正しい公休数（土日祝の数）をカレンダーから自動計算する
                    target_off = sum(1 for d in dates if is_holiday_jp(d))

                    # 罫線・縮小表示・背景色・文字サイズの一括適用
                    for row_idx, row in enumerate(ws.iter_rows()):
                        is_empty_row = all(not cell.value for cell in row)
                        for cell in row:
                            if row_idx == 0:
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            else:
                                # 🌟【進化】全体の基本文字サイズを「9」に統一
                                cell.font = Font(size=9)
                                # 縮小して全体を表示
                                cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
                                
                                # 部署間の空白行以外に、格子罫線と背景色を適用（オシャレ仕様）
                                if not is_empty_row:
                                    cell.border = thin_border
                                    
                                    # 2行目以降、かつ日付の列範囲であれば土日祝の背景色を塗る
                                    if row_idx >= 1 and 3 <= cell.column <= len(dates) + 2:
                                        d = dates[cell.column - 3]
                                        if d.weekday() == 6 or jpholiday.is_holiday(d):
                                            cell.fill = hol_fill
                                        elif d.weekday() == 5:
                                            cell.fill = sat_fill

                    # 1行目の結合と文字入力（見出しタイトル）
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(weekday_row))
                    ws.cell(row=1, column=1).value = f"【 {era_str} 】　　　　　　　　　　　　　　　　　OFF {target_off}回"
                    # 🌟【進化】タイトルだけバランスを見てサイズ「11」に指定
                    ws.cell(row=1, column=1).font = Font(bold=True, size=11)

                    # 土日の見出し文字の色付け（サイズ9を維持しつつ太字＋色付け）
                    for i, d in enumerate(dates):
                        col_idx = i + 3
                        color = None
                        if d.weekday() == 6 or jpholiday.is_holiday(d): color = "FF0000"
                        elif d.weekday() == 5: color = "0000FF"
                        
                        if color:
                            ws.cell(row=2, column=col_idx).font = Font(color=color, bold=True, size=9)
                            ws.cell(row=3, column=col_idx).font = Font(color=color, bold=True, size=9)

                    # 赤・青希望の文字色付け（サイズ9を維持しつつ太字＋色付け）
                    for (r_i, c_i), hex_color in cell_styles.items():
                        ws.cell(row=r_i+1, column=c_i+1).font = Font(color=hex_color, bold=True, size=9)

                    # 列幅の調整
                    ws.column_dimensions['A'].width = 8
                    ws.column_dimensions['B'].width = 14
                    import openpyxl.utils
                    for i in range(len(dates)):
                        col_letter = openpyxl.utils.get_column_letter(i + 3)
                        ws.column_dimensions[col_letter].width = 4.2
                    # サマリー列の幅をスリム化（3.5に設定）
                    for i in range(6):
                        col_letter = openpyxl.utils.get_column_letter(len(dates) + 3 + i)
                        ws.column_dimensions[col_letter].width = 3.5

                    # 印刷設定 (横向き・1ページに収める・余白最小化)
                    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                    ws.page_setup.fitToPage = True
                    ws.page_setup.fitToHeight = 1
                    ws.page_setup.fitToWidth = 1
                    
                    ws.page_margins.left = 0.2
                    ws.page_margins.right = 0.2
                    ws.page_margins.top = 0.5
                    ws.page_margins.bottom = 0.5
                    ws.page_margins.header = 0.2
                    ws.page_margins.footer = 0.2

                st.download_button(
                    label="📊 Excel形式でダウンロード (最終形態フォーマット)",
                    data=output.getvalue(),
                    file_name=f"shift_{selected_year}_{selected_month}_final.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )